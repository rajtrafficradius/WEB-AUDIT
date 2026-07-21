"""Professional XLSX workbook renderers for the audit deliverable package.

Pure ``dict -> files`` renderers built on openpyxl. No Django imports: the
input is the compiled run-data dictionary produced by
``exporters.run_data.compile_run_data`` and the output is the V18-shaped
workbook tree of the client package.

Every sheet follows the house convention: a brand band (rows 1-2), a styled
header row (row 4), zebra-striped data from row 5, frozen panes at A5, an
auto-filter across the header + data range, explicit column widths, and a
closing Methodology sheet where one makes sense.

Evidence-first: no metric is invented. Values the pipeline did not measure are
rendered as ``Unavailable`` (with the provider reason when one exists), empty
sections carry an explicit professional statement instead of a blank grid, and
an on-page proposal that matches the current value is labelled
``No change required`` rather than shipped as fake work.
"""

from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path
from typing import Any
from urllib.parse import urlsplit

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from exporters import paths as tree
from exporters.brand import (
    ACCENT,
    CRITICAL,
    INK,
    ISSUE_FILL,
    PAPER,
    RULE,
    SEVERITY_FILL,
    WHITE,
    argb,
)

UNAVAILABLE = "Unavailable"
NO_CHANGE = "No change required"
EMPTY_MESSAGE = "No issues detected in this category during the crawl window."
KEYWORD_VOLUME_NOTE = (
    "Note: keyword search volumes and rankings require an approved GSC or SEMrush "
    "connection. Only crawl-observed topics are listed; no volumes are estimated."
)

# Column width presets by content type.
W_URL = 55
W_TEXT = 45
W_NUM = 12
W_DATE = 14
W_LABEL = 18
W_WEEK = 5

SEVERITY_COLORS = SEVERITY_FILL
ISSUE_COLORS = ISSUE_FILL
COVERAGE_COLORS: dict[str, tuple[str, str]] = {
    "covered": ISSUE_FILL["ok"],
    "partial": ISSUE_FILL["review"],
    "gap": ISSUE_FILL["missing"],
}

MONEY_PAGE_TYPES = frozenset({"Homepage", "Product", "Collection"})
ANSWER_SCHEMA = frozenset({"faqpage", "qapage", "howto", "question", "answer"})
LOCAL_SCHEMA = frozenset(
    {"localbusiness", "postaladdress", "place", "organization", "store"}
)

TITLE_MIN, TITLE_MAX = 30, 60
META_MIN, META_MAX = 70, 160
THIN_WORDS = 250
# A revenue page below this is genuinely broken; between the two bounds is
# normal for ecommerce product templates and is NOT reported as a defect.
THIN_WORDS_CRITICAL = 120
# Homepages and collection pages are expected to carry more copy.
THIN_WORDS_CONTEXT = 200
SLOW_MS = 1500
# When one signal fires on more pages than this, the findings sheet shows a
# single aggregate row instead of a wall of identical rows.
SIGNAL_AGGREGATE_THRESHOLD = 8


def _argb(color: str) -> str:
    return argb(color)


def _fill(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", start_color=_argb(color), end_color=_argb(color))


_THIN_RULE = Side(style="thin", color=_argb(RULE)[2:])
_BORDER = Border(left=_THIN_RULE, right=_THIN_RULE, top=_THIN_RULE, bottom=_THIN_RULE)


def _guard(value: str) -> str:
    """Formula-injection guard: neutralise strings Excel could execute."""
    if value[:1] in ("=", "+", "-", "@"):
        return "'" + value
    return value


def _cell_value(value: Any) -> Any:
    """Coerce a raw value for a worksheet cell. Numbers stay numbers."""
    if value is None:
        return UNAVAILABLE
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, int | float):
        return value
    if isinstance(value, list | tuple):
        joined = ", ".join(str(item) for item in value)
        return _guard(joined) if joined else ""
    text = str(value).strip()
    if not text:
        return ""
    return _guard(text)


def _severity_token(value: Any) -> str:
    return str(value or "").strip().casefold()


def _add_sheet(
    workbook: Workbook,
    meta: dict[str, str],
    name: str,
    headers: list[str],
    rows: list[list[Any]],
    widths: list[int],
    *,
    wrap_columns: frozenset[int] = frozenset(),
    color_column: int | None = None,
    color_map: dict[str, tuple[str, str]] | None = None,
    empty_message: str = EMPTY_MESSAGE,
    note: str | None = None,
) -> Worksheet:
    """Write one branded sheet: band rows 1-2, header row 4, data from row 5."""
    sheet = workbook.create_sheet(title=name[:31])
    column_count = len(headers)
    last_letter = get_column_letter(column_count)

    # Brand band.
    sheet.merge_cells(f"A1:{last_letter}1")
    sheet.merge_cells(f"A2:{last_letter}2")
    band_fill = _fill(INK)
    for column in range(1, column_count + 1):
        for row in (1, 2):
            cell = sheet.cell(row=row, column=column)
            cell.fill = band_fill
    title_cell = sheet["A1"]
    title_cell.value = _guard(f"{meta['client']} — {meta['title']}")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=_argb(WHITE)[2:])
    title_cell.alignment = Alignment(vertical="center")
    subtitle_cell = sheet["A2"]
    subtitle_cell.value = _guard(
        f"Prepared by Traffic Radius · Evidence as of {meta['as_of']} · Run {meta['run_id']}"
    )
    subtitle_cell.font = Font(name="Calibri", italic=True, size=9, color=_argb(RULE)[2:])
    subtitle_cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 15

    # Header row 4.
    header_fill = _fill(ACCENT)
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=4, column=index, value=_guard(header))
        cell.font = Font(name="Calibri", bold=True, size=11, color=_argb(WHITE)[2:])
        cell.fill = header_fill
        cell.border = _BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[4].height = 24

    body_rows = rows or [[empty_message] + [""] * (column_count - 1)]
    zebra_fill = _fill(PAPER)
    for row_offset, row_values in enumerate(body_rows):
        row_number = 5 + row_offset
        for column_index in range(1, column_count + 1):
            raw = row_values[column_index - 1] if column_index <= len(row_values) else ""
            cell = sheet.cell(row=row_number, column=column_index, value=_cell_value(raw))
            cell.border = _BORDER
            font_color = _argb(INK)[2:]
            fill = zebra_fill if row_number % 2 == 0 else None
            if color_column is not None and column_index == color_column and color_map:
                tone = color_map.get(_severity_token(raw))
                if tone is not None:
                    fill = _fill(tone[0])
                    font_color = _argb(tone[1])[2:]
            if fill is not None:
                cell.fill = fill
            cell.font = Font(name="Calibri", size=10, color=font_color)
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(column_index in wrap_columns),
            )

    last_data_row = 4 + len(body_rows)
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:{last_letter}{last_data_row}"
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    if note:
        note_row = last_data_row + 2
        sheet.merge_cells(f"A{note_row}:{last_letter}{note_row}")
        note_cell = sheet.cell(row=note_row, column=1, value=_guard(note))
        note_cell.font = Font(name="Calibri", bold=True, size=10, color=_argb(CRITICAL)[2:])
        note_cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.row_dimensions[note_row].height = 30
    return sheet


def _methodology_rows(data: dict[str, Any], measured: str) -> list[list[Any]]:
    run = data.get("run", {})
    coverage = run.get("evidence_coverage")
    coverage_text = f"{coverage * 100:.0f}%" if isinstance(coverage, int | float) else UNAVAILABLE
    integrity = data.get("crawl_integrity") or {}
    rows: list[list[Any]] = [
        ["What was measured", measured],
        ["Evidence cutoff", run.get("evidence_as_of") or UNAVAILABLE],
        ["Ruleset version", run.get("rule_version") or UNAVAILABLE],
        ["Run", run.get("id") or UNAVAILABLE],
        ["Client domain", data.get("client", {}).get("domain") or UNAVAILABLE],
        ["Configured page budget", run.get("configured_page_budget")],
        ["Weighted evidence coverage", coverage_text],
        [
            "Coverage interpretation",
            run.get("coverage_interpretation") or UNAVAILABLE,
        ],
        [
            "Evidence policy",
            "Canonical crawl evidence only. Unavailable fields are labelled explicitly; "
            "no substitute metrics are invented.",
        ],
    ]
    if integrity:
        rows.append(["Crawl integrity", integrity.get("status") or UNAVAILABLE])
        rows.append(["Pages fetched", integrity.get("fetched_pages")])
        rows.append(["Challenged pages", integrity.get("challenged_pages")])
        if integrity.get("note"):
            rows.append(["Crawl integrity note", integrity.get("note")])
    for limitation in data.get("limitations") or []:
        rows.append(["Limitation", limitation])
    return rows


def _add_methodology(
    workbook: Workbook,
    meta: dict[str, str],
    data: dict[str, Any],
    measured: str,
    *,
    extra: list[list[Any]] | None = None,
) -> None:
    rows = _methodology_rows(data, measured)
    rows.extend(extra or [])
    _add_sheet(
        workbook,
        meta,
        "Methodology",
        ["Field", "Detail"],
        rows,
        [26, 90],
        wrap_columns=frozenset({2}),
    )


def _meta(data: dict[str, Any], title: str) -> dict[str, str]:
    run = data.get("run", {})
    return {
        "client": str(data.get("client", {}).get("name") or "Client"),
        "title": title,
        "as_of": str(run.get("evidence_as_of") or UNAVAILABLE),
        "run_id": str(run.get("id") or UNAVAILABLE),
    }


def _new_workbook() -> Workbook:
    workbook = Workbook()
    default = workbook.active
    if default is not None:
        workbook.remove(default)
    return workbook


def _save(workbook: Workbook, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def _findings_for(data: dict[str, Any], *keys: str) -> list[dict[str, Any]]:
    wanted = {key.casefold() for key in keys}
    return [
        finding
        for finding in data.get("findings") or []
        if str(finding.get("category") or "").casefold() in wanted
    ]


def _finding_rows(findings: list[dict[str, Any]]) -> list[list[Any]]:
    return [
        [
            finding.get("id"),
            finding.get("priority"),
            finding.get("severity"),
            finding.get("title"),
            finding.get("description"),
            finding.get("impact"),
            finding.get("affected_count"),
            finding.get("confidence"),
            finding.get("effort"),
            finding.get("evidence_ids"),
        ]
        for finding in findings
    ]


FINDING_HEADERS = [
    "Finding ID", "Priority", "Severity", "Title", "Description",
    "Impact", "Affected Count", "Confidence", "Effort", "Evidence IDs",
]
FINDING_WIDTHS = [W_NUM, W_NUM, W_NUM, W_TEXT, W_TEXT, W_TEXT, W_NUM, W_NUM, W_NUM, 24]


# --------------------------------------------------------------------------- accessors


def _pages(data: dict[str, Any]) -> list[dict[str, Any]]:
    return list(data.get("pages") or [])


def _keywords(data: dict[str, Any]) -> list[dict[str, Any]]:
    return [row for row in (data.get("keywords") or []) if isinstance(row, dict)]


def _clusters(data: dict[str, Any]) -> list[dict[str, Any]]:
    return [row for row in (data.get("keyword_clusters") or []) if isinstance(row, dict)]


def _competitors(data: dict[str, Any]) -> list[dict[str, Any]]:
    return [row for row in (data.get("competitors") or []) if isinstance(row, dict)]


def _market(data: dict[str, Any]) -> dict[str, Any]:
    market = data.get("market")
    return market if isinstance(market, dict) else {}


def _backlinks(data: dict[str, Any]) -> dict[str, Any]:
    backlinks = data.get("backlinks")
    return backlinks if isinstance(backlinks, dict) else {}


def _backlinks_available(data: dict[str, Any]) -> bool:
    return str(_backlinks(data).get("status") or "").casefold() == "available"


def _unavailable_reason(section: dict[str, Any], fallback: str) -> str:
    reason = section.get("unavailable_reason")
    text = str(reason).strip() if reason else ""
    return text or fallback


def _unavailable_message(section: dict[str, Any], fallback: str) -> str:
    return f"{UNAVAILABLE} — {_unavailable_reason(section, fallback)}"


def _facts(page: dict[str, Any]) -> dict[str, Any]:
    facts = page.get("_facts")
    return facts if isinstance(facts, dict) else {}


def _business_profile(data: dict[str, Any]) -> str:
    project = data.get("project") or {}
    return str(project.get("business_profile") or "").casefold()


def _is_ecommerce(data: dict[str, Any]) -> bool:
    profile = _business_profile(data)
    if "ecom" in profile or "hybrid" in profile or "retail" in profile:
        return True
    return any(
        str(page.get("page_type")) in {"Product", "Collection"} for page in _pages(data)
    )


def _is_local(data: dict[str, Any]) -> bool:
    profile = _business_profile(data)
    if "local" in profile or "hybrid" in profile or "service" in profile:
        return True
    for page in _pages(data):
        for schema in page.get("schema_types") or []:
            if str(schema).casefold() in LOCAL_SCHEMA:
                return True
    return False


# --------------------------------------------------------------------------- page rules


def _canonical_state(page: dict[str, Any]) -> str:
    canonical = page.get("canonical_url")
    if not canonical:
        return "Missing"
    normalized = str(page.get("normalized_url") or "")
    if str(canonical).rstrip("/").casefold() == normalized.rstrip("/").casefold():
        return "Match"
    return "Mismatch"


def _url_depth(url: str) -> int:
    return len([segment for segment in urlsplit(url).path.split("/") if segment])


def _text_length(value: Any) -> int | None:
    if value is None:
        return None
    return len(str(value))


def _title_issue(page: dict[str, Any]) -> str:
    title = page.get("title")
    if not title:
        return "Missing"
    length = len(str(title))
    if length > TITLE_MAX:
        return "Too long"
    if length < TITLE_MIN:
        return "Too short"
    return "OK"


def _meta_issue(page: dict[str, Any]) -> str:
    meta = page.get("meta_description")
    if not meta:
        return "Missing"
    length = len(str(meta))
    if length > META_MAX:
        return "Too long"
    if length < META_MIN:
        return "Too short"
    return "OK"


def _h1_count(page: dict[str, Any]) -> int:
    values = _facts(page).get("h1_values")
    if isinstance(values, list):
        return len([value for value in values if str(value or "").strip()])
    h1 = page.get("h1")
    if not h1:
        return 0
    return len([part for part in str(h1).split(" | ") if part.strip()])


def _h1_issue(page: dict[str, Any]) -> str:
    count = _h1_count(page)
    if count == 0:
        return "Missing"
    if count > 1:
        return "Multiple captured"
    return "OK"


def _indexability_blocked(page: dict[str, Any]) -> bool:
    value = page.get("indexability")
    if value is False:
        return True
    text = str(value or "").casefold()
    return "noindex" in text and "no noindex" not in text


def _schema_text(page: dict[str, Any]) -> str:
    types = page.get("schema_types") or []
    return ", ".join(str(item) for item in types) if types else "None detected"


def _has_schema(page: dict[str, Any], needle: str) -> bool:
    return any(needle in str(item).casefold() for item in (page.get("schema_types") or []))


EXPECTED_SCHEMA: dict[str, str] = {
    "Homepage": "Organization / WebSite",
    "Product": "Product",
    "Collection": "ItemList / CollectionPage",
    "Editorial": "Article / BlogPosting",
    "Information": "WebPage",
    "Utility": "WebPage",
    "Other": "WebPage",
}


def _schema_gap(page: dict[str, Any]) -> str:
    """State the structured-data gap for a page without asserting anything unmeasured."""
    expected = EXPECTED_SCHEMA.get(str(page.get("page_type") or ""), "WebPage")
    declared = {str(item).casefold() for item in (page.get("schema_types") or [])}
    if not declared:
        return f"No structured data detected; {expected} is expected for this page type"
    for candidate in expected.split(" / "):
        if candidate.casefold() in declared:
            return "Expected type present"
    return f"Declared types do not include {expected}"


def _answer_ready(page: dict[str, Any]) -> tuple[int, str]:
    """Deterministic answer-readiness score from crawl facts only (0-100)."""
    schema_types = {str(item).casefold() for item in (page.get("schema_types") or [])}
    parts: list[str] = []
    score = 0
    if schema_types:
        score += 30
        parts.append("structured data present (+30)")
    if schema_types & ANSWER_SCHEMA:
        score += 20
        parts.append("answer schema present (+20)")
    if _h1_count(page) == 1:
        score += 20
        parts.append("single H1 (+20)")
    word_count = page.get("word_count")
    if isinstance(word_count, int) and word_count >= 300:
        score += 15
        parts.append("300+ words (+15)")
    if page.get("meta_description"):
        score += 15
        parts.append("meta description present (+15)")
    return score, "; ".join(parts) or "No answer-readiness signals observed"


# --------------------------------------------------------------------------- proposals


def _proposal_state(current: Any, proposed: Any) -> tuple[Any, str]:
    """Return (rendered proposal, status). Identical proposals are never shipped."""
    current_text = str(current or "").strip()
    proposed_text = str(proposed or "").strip()
    if not proposed_text:
        return (UNAVAILABLE, "No proposal generated")
    if proposed_text.casefold() == current_text.casefold():
        return (NO_CHANGE, NO_CHANGE)
    return (proposed_text, "Proposed")


def _proposals(data: dict[str, Any]) -> list[dict[str, Any]]:
    """Return on-page proposals, falling back to the legacy metadata review."""
    rows = data.get("onpage_proposals")
    if isinstance(rows, list) and rows:
        return [row for row in rows if isinstance(row, dict)]
    legacy: list[dict[str, Any]] = []
    deployment = data.get("deployment") or {}
    for entry in deployment.get("metadata_review") or []:
        legacy.append(
            {
                "page_id": entry.get("page_id"),
                "url": entry.get("url"),
                "page_type": entry.get("page_type"),
                "current_title": entry.get("current_title"),
                "proposed_title": entry.get("proposed_title"),
                "title_rationale": entry.get("title_issue"),
                "current_meta": entry.get("current_meta_description"),
                "proposed_meta": entry.get("proposed_meta_description"),
                "meta_rationale": entry.get("meta_description_issue"),
                "current_h1": entry.get("current_h1"),
                "proposed_h1": entry.get("proposed_h1"),
                "h1_rationale": entry.get("h1_issue"),
                "target_keyword": entry.get("target_keyword"),
                "target_volume": None,
                "source": "deterministic",
                "approval_status": entry.get("approval_status")
                or "withheld_pending_editorial_review",
                "evidence_ids": [entry.get("evidence_id")] if entry.get("evidence_id") else [],
            }
        )
    return legacy


# --------------------------------------------------------------------------- technical


TECHNICAL_INVENTORY_HEADERS = [
    "URL", "Status Code", "Title", "Title Length", "Title Issue",
    "Meta Description", "Meta Length", "Meta Issue", "H1", "H1 Count", "H1 Issue",
    "Word Count", "Canonical", "Canonical State", "Indexability", "Schema Types",
    "Internal Links", "External Links", "Images Total", "Images Missing Alt",
    "Response (ms)", "Body Bytes", "Page Type", "Depth", "Evidence ID",
]
TECHNICAL_INVENTORY_WIDTHS = [
    W_URL, W_NUM, W_TEXT, W_NUM, W_LABEL, W_TEXT, W_NUM, W_LABEL, W_TEXT, W_NUM,
    W_LABEL, W_NUM, W_URL, W_LABEL, 28, 26, W_NUM, W_NUM, W_NUM, W_NUM, W_NUM,
    W_NUM, W_LABEL, W_NUM, W_NUM,
]


def _inventory_row(page: dict[str, Any]) -> list[Any]:
    return [
        page.get("normalized_url"),
        page.get("status_code"),
        page.get("title"),
        _text_length(page.get("title")),
        _title_issue(page),
        page.get("meta_description"),
        _text_length(page.get("meta_description")),
        _meta_issue(page),
        page.get("h1"),
        _h1_count(page),
        _h1_issue(page),
        page.get("word_count"),
        page.get("canonical_url"),
        _canonical_state(page),
        page.get("indexability"),
        _schema_text(page),
        page.get("internal_links"),
        page.get("external_links"),
        page.get("images_total"),
        page.get("images_missing_alt"),
        page.get("response_ms"),
        page.get("body_bytes"),
        page.get("page_type"),
        _url_depth(str(page.get("normalized_url") or "")),
        page.get("evidence_id"),
    ]


def _technical_audit(data: dict[str, Any], path: Path) -> Path:
    meta = _meta(data, "Technical Audit Report")
    pages = _pages(data)
    workbook = _new_workbook()

    _add_sheet(
        workbook, meta, "Full Site Inventory",
        TECHNICAL_INVENTORY_HEADERS,
        [_inventory_row(page) for page in pages],
        TECHNICAL_INVENTORY_WIDTHS,
        wrap_columns=frozenset({3, 6, 9}),
        empty_message="No pages were captured during the crawl window.",
    )

    errors = [
        [page.get("normalized_url"), page.get("status_code"), page.get("page_type"),
         page.get("indexability"), page.get("internal_links"), page.get("evidence_id")]
        for page in pages
        if isinstance(page.get("status_code"), int) and page["status_code"] >= 400
    ]
    _add_sheet(
        workbook, meta, "Error Pages",
        ["URL", "Status Code", "Page Type", "Indexability", "Internal Links",
         "Evidence ID"],
        errors, [W_URL, W_NUM, W_LABEL, 28, W_NUM, W_NUM],
    )

    redirects = [
        [page.get("normalized_url"), page.get("status_code"),
         " → ".join(str(hop) for hop in page.get("redirect_chain") or []),
         len(page.get("redirect_chain") or []), page.get("page_type"),
         page.get("evidence_id")]
        for page in pages
        if len(page.get("redirect_chain") or []) > 1
    ]
    _add_sheet(
        workbook, meta, "Redirects",
        ["URL", "Status Code", "Redirect Chain", "Hops", "Page Type", "Evidence ID"],
        redirects, [W_URL, W_NUM, W_URL + 20, W_NUM, W_LABEL, W_NUM],
        wrap_columns=frozenset({3}),
    )

    canonical_issues = [
        [page.get("normalized_url"), page.get("canonical_url"), _canonical_state(page),
         page.get("page_type"), page.get("evidence_id")]
        for page in pages
        if _canonical_state(page) != "Match"
    ]
    _add_sheet(
        workbook, meta, "Canonical Issues",
        ["URL", "Declared Canonical", "State", "Page Type", "Evidence ID"],
        canonical_issues, [W_URL, W_URL, W_NUM, W_LABEL, W_NUM],
        color_column=3, color_map=ISSUE_COLORS,
    )

    hash_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for page in pages:
        digest = page.get("body_sha256")
        if digest:
            hash_groups[str(digest)].append(page)
    duplicates: list[list[Any]] = []
    for digest in sorted(hash_groups):
        group = hash_groups[digest]
        if len(group) > 1:
            for page in group:
                duplicates.append(
                    [page.get("normalized_url"), digest, len(group),
                     page.get("page_type"), page.get("evidence_id")]
                )
    _add_sheet(
        workbook, meta, "Duplicate Content",
        ["URL", "Body SHA-256", "Group Size", "Page Type", "Evidence ID"],
        duplicates, [W_URL, 40, W_NUM, W_LABEL, W_NUM],
    )

    indexability = [
        [page.get("normalized_url"), page.get("indexability"), page.get("status_code"),
         page.get("page_type"), page.get("evidence_id")]
        for page in pages
        if _indexability_blocked(page)
    ]
    _add_sheet(
        workbook, meta, "Indexability",
        ["URL", "Indexability Signal", "Status Code", "Page Type", "Evidence ID"],
        indexability, [W_URL, 34, W_NUM, W_LABEL, W_NUM],
        empty_message="No noindex or robots exclusions were observed during the crawl window.",
    )

    _add_sheet(
        workbook, meta, "Title Issues",
        ["URL", "Page Type", "Current Title", "Length", "Issue", "Evidence ID"],
        [
            [page.get("normalized_url"), page.get("page_type"), page.get("title"),
             _text_length(page.get("title")), _title_issue(page), page.get("evidence_id")]
            for page in pages
            if _title_issue(page) != "OK"
        ],
        [W_URL, W_LABEL, W_TEXT, W_NUM, W_LABEL, W_NUM],
        wrap_columns=frozenset({3}), color_column=5, color_map=ISSUE_COLORS,
        note=f"Length rule: {TITLE_MIN}-{TITLE_MAX} characters observed in the crawled HTML.",
    )

    _add_sheet(
        workbook, meta, "Meta Issues",
        ["URL", "Page Type", "Current Meta Description", "Length", "Issue",
         "Evidence ID"],
        [
            [page.get("normalized_url"), page.get("page_type"),
             page.get("meta_description"), _text_length(page.get("meta_description")),
             _meta_issue(page), page.get("evidence_id")]
            for page in pages
            if _meta_issue(page) != "OK"
        ],
        [W_URL, W_LABEL, W_TEXT, W_NUM, W_LABEL, W_NUM],
        wrap_columns=frozenset({3}), color_column=5, color_map=ISSUE_COLORS,
        note=f"Length rule: {META_MIN}-{META_MAX} characters observed in the crawled HTML.",
    )

    _add_sheet(
        workbook, meta, "H1 Issues",
        ["URL", "Page Type", "Current H1", "H1 Count", "Issue", "Evidence ID"],
        [
            [page.get("normalized_url"), page.get("page_type"), page.get("h1"),
             _h1_count(page), _h1_issue(page), page.get("evidence_id")]
            for page in pages
            if _h1_issue(page) != "OK"
        ],
        [W_URL, W_LABEL, W_TEXT, W_NUM, W_LABEL, W_NUM],
        wrap_columns=frozenset({3}), color_column=5, color_map=ISSUE_COLORS,
    )

    alt_pages = [page for page in pages if isinstance(page.get("images_missing_alt"), int)]
    _add_sheet(
        workbook, meta, "Image Alt Issues",
        ["URL", "Images Total", "Images Missing Alt", "Page Type", "Evidence ID"],
        [
            [page.get("normalized_url"), page.get("images_total"),
             page.get("images_missing_alt"), page.get("page_type"),
             page.get("evidence_id")]
            for page in alt_pages
            if (page.get("images_missing_alt") or 0) > 0
        ],
        [W_URL, W_NUM, W_NUM, W_LABEL, W_NUM],
        empty_message=(
            EMPTY_MESSAGE if alt_pages
            else "Image inventories were unavailable for this run; alt coverage was not assessed."
        ),
    )

    _add_sheet(
        workbook, meta, "Structured Data",
        ["URL", "Page Type", "Declared Schema Types", "Types Detected",
         "Expected For Page Type", "Gap", "Evidence ID"],
        [
            [
                page.get("normalized_url"), page.get("page_type"), _schema_text(page),
                len(page.get("schema_types") or []),
                EXPECTED_SCHEMA.get(str(page.get("page_type") or ""), "WebPage"),
                _schema_gap(page), page.get("evidence_id"),
            ]
            for page in pages
        ],
        [W_URL, W_LABEL, 30, W_NUM, 26, W_TEXT, W_NUM],
        wrap_columns=frozenset({3, 6}),
        empty_message="No pages were captured during the crawl window.",
        note=(
            "Expected types are the schema.org types a page of this classification "
            "normally declares. Only types observed in the crawled HTML are reported."
        ),
    )

    _add_sheet(
        workbook, meta, "Findings Register",
        FINDING_HEADERS,
        _finding_rows(_findings_for(data, "technical", "on_page", "onpage", "on-page")),
        FINDING_WIDTHS, wrap_columns=frozenset({4, 5, 6}),
        color_column=3, color_map=SEVERITY_COLORS,
    )
    _add_methodology(
        workbook, meta, data,
        "HTTP status, redirect chains, canonical declarations, duplicate body hashes, "
        "indexability signals and metadata quality for every page fetched in the "
        "approved-domain crawl.",
    )
    return _save(workbook, path)


# --------------------------------------------------------------------------- content


def _keyword_coverage(data: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    coverage: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for keyword in _keywords(data):
        url = str(keyword.get("landing_url") or "").strip()
        if url:
            coverage[url].append(keyword)
    return coverage


def _cannibalisation_rows(data: dict[str, Any]) -> list[list[Any]]:
    pages = _pages(data)
    rows: list[list[Any]] = []

    title_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for page in pages:
        title = str(page.get("title") or "").strip().casefold()
        if title:
            title_groups[title].append(page)
    for title in sorted(title_groups):
        group = title_groups[title]
        if len(group) < 2:
            continue
        rows.append([
            "Duplicate title",
            group[0].get("title"),
            len(group),
            ", ".join(str(page.get("normalized_url")) for page in group[:6]),
            "Differentiate titles or consolidate the weaker page into the stronger one.",
            [page.get("evidence_id") for page in group[:6]],
        ])

    keyword_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for keyword in _keywords(data):
        phrase = str(keyword.get("phrase") or "").strip().casefold()
        url = str(keyword.get("landing_url") or "").strip()
        if phrase and url:
            keyword_groups[phrase].append(keyword)
    for phrase in sorted(keyword_groups):
        group = keyword_groups[phrase]
        urls = {str(item.get("landing_url")) for item in group}
        if len(urls) < 2:
            continue
        rows.append([
            "Multiple URLs for one keyword",
            group[0].get("phrase"),
            len(urls),
            ", ".join(sorted(urls)[:6]),
            "Pick one canonical target and internally link the alternates to it.",
            sorted({eid for item in group for eid in (item.get("evidence_ids") or [])}),
        ])
    return rows


def _content_audit(data: dict[str, Any], path: Path) -> Path:
    meta = _meta(data, "Content Audit Workbook")
    pages = _pages(data)
    keywords = _keywords(data)
    coverage = _keyword_coverage(data)
    market = _market(data)
    workbook = _new_workbook()

    _add_sheet(
        workbook, meta, "Full Page Inventory",
        ["URL", "Page Type", "Title", "H1", "Word Count", "Mapped Keywords",
         "Internal Links", "Images", "Schema Types", "Status", "Evidence ID"],
        [
            [
                page.get("normalized_url"), page.get("page_type"), page.get("title"),
                page.get("h1"), page.get("word_count"),
                len(coverage.get(str(page.get("normalized_url") or ""), [])),
                page.get("internal_links"), page.get("images_total"),
                _schema_text(page), page.get("status_code"), page.get("evidence_id"),
            ]
            for page in pages
        ],
        [W_URL, W_LABEL, W_TEXT, W_TEXT, W_NUM, W_NUM, W_NUM, W_NUM, 26, W_NUM, W_NUM],
        wrap_columns=frozenset({3, 4}),
        empty_message="No pages were captured during the crawl window.",
    )

    _add_sheet(
        workbook, meta, "Content Gap",
        ["Keyword ID", "Phrase", "Volume", "CPC", "Competition", "Intent",
         "Funnel Stage", "Cluster", "Opportunity", "Source", "Evidence IDs"],
        [
            [
                keyword.get("id"), keyword.get("phrase"), keyword.get("search_volume"),
                keyword.get("cpc"), keyword.get("competition"), keyword.get("intent"),
                keyword.get("funnel_stage"), keyword.get("cluster"),
                keyword.get("opportunity"), keyword.get("source"),
                keyword.get("evidence_ids"),
            ]
            for keyword in keywords
            if not str(keyword.get("landing_url") or "").strip()
        ],
        [W_NUM, W_TEXT, W_NUM, W_NUM, W_NUM, W_LABEL, W_LABEL, 24, W_TEXT, W_LABEL, 22],
        wrap_columns=frozenset({2, 9}),
        empty_message=(
            "Every measured keyword already maps to a crawled URL."
            if keywords
            else _unavailable_message(market, "no keyword provider connected")
        ),
        note=None if keywords else KEYWORD_VOLUME_NOTE,
    )

    coverage_rows: list[list[Any]] = []
    for page in pages:
        url = str(page.get("normalized_url") or "")
        mapped = coverage.get(url, [])
        if not mapped:
            continue
        volumes = [k.get("search_volume") for k in mapped
                   if isinstance(k.get("search_volume"), int | float)]
        best = min(
            (k for k in mapped if isinstance(k.get("position"), int)),
            key=lambda item: item["position"],
            default=None,
        )
        coverage_rows.append([
            url, page.get("page_type"), len(mapped),
            ", ".join(str(k.get("phrase")) for k in mapped[:6]),
            sum(volumes) if volumes else None,
            (best or {}).get("phrase"), (best or {}).get("position"),
            page.get("word_count"), page.get("evidence_id"),
        ])
    _add_sheet(
        workbook, meta, "Keyword Coverage by URL",
        ["URL", "Page Type", "Keywords Mapped", "Sample Keywords", "Total Volume",
         "Best Ranking Keyword", "Best Position", "Word Count", "Evidence ID"],
        coverage_rows,
        [W_URL, W_LABEL, W_NUM, W_TEXT, W_NUM, W_TEXT, W_NUM, W_NUM, W_NUM],
        wrap_columns=frozenset({4}),
        empty_message=(
            _unavailable_message(market, "no keyword provider connected")
            if not keywords
            else "No measured keyword mapped to a crawled URL."
        ),
    )

    counted = [page for page in pages if isinstance(page.get("word_count"), int)]
    thin_rows: list[list[Any]] = [
        [page.get("normalized_url"), page.get("word_count"), page.get("page_type"),
         "Money page" if str(page.get("page_type")) in MONEY_PAGE_TYPES else "Supporting",
         page.get("title"), page.get("evidence_id")]
        for page in sorted(counted, key=lambda item: item.get("word_count") or 0)
        if (page.get("word_count") or 0) < THIN_WORDS
    ]
    duplicate_titles: Counter[str] = Counter(
        str(page.get("title")).strip().casefold() for page in pages if page.get("title")
    )
    for page in pages:
        title = str(page.get("title") or "").strip().casefold()
        if title and duplicate_titles[title] > 1:
            thin_rows.append([
                page.get("normalized_url"), page.get("word_count"), page.get("page_type"),
                "Duplicate title", page.get("title"), page.get("evidence_id"),
            ])
    _add_sheet(
        workbook, meta, "Duplicate or Thin Content",
        ["URL", "Word Count", "Page Type", "Signal", "Title", "Evidence ID"],
        thin_rows, [W_URL, W_NUM, W_LABEL, W_LABEL, W_TEXT, W_NUM],
        wrap_columns=frozenset({5}),
        empty_message=(
            EMPTY_MESSAGE if counted
            else "Word counts were unavailable for this run; thin content was not assessed."
        ),
        note=f"Threshold: pages under {THIN_WORDS} observed body words are flagged.",
    )

    depth_rows: list[list[Any]] = []
    if counted:
        counts = sorted(int(page["word_count"]) for page in counted)
        for label, low, high in (
            ("Under 250 words", 0, 249), ("250-599 words", 250, 599),
            ("600-1199 words", 600, 1199), ("1200-2399 words", 1200, 2399),
            ("2400+ words", 2400, 10**9),
        ):
            matched = [value for value in counts if low <= value <= high]
            depth_rows.append([
                label, len(matched), round(len(matched) / len(counts), 3),
                sum(matched) if matched else None,
            ])
        depth_rows.append(["Median word count", counts[len(counts) // 2], None, None])
        depth_rows.append(["Shortest page (words)", counts[0], None, None])
        depth_rows.append(["Longest page (words)", counts[-1], None, None])
        depth_rows.append([
            "Pages with no measured word count", len(pages) - len(counted), None, None,
        ])
    _add_sheet(
        workbook, meta, "Content Depth Distribution",
        ["Band", "Pages", "Share", "Total Words"],
        depth_rows, [26, W_NUM, W_NUM, W_NUM],
        empty_message=(
            f"{UNAVAILABLE} — word counts were not captured for this run."
        ),
    )

    _add_sheet(
        workbook, meta, "Cannibalisation Signals",
        ["Signal", "Shared Value", "Pages", "URLs", "Recommended Resolution",
         "Evidence IDs"],
        _cannibalisation_rows(data), [W_LABEL, W_TEXT, W_NUM, W_URL + 20, W_TEXT, 22],
        wrap_columns=frozenset({2, 4, 5}),
        empty_message="No overlapping targeting signals were observed in the crawl.",
    )
    _add_methodology(
        workbook, meta, data,
        "Page-level content inventory, measured keyword coverage per URL, thin and "
        "duplicate content signals, and overlapping targeting observed across the "
        "approved-domain crawl.",
    )
    return _save(workbook, path)


# --------------------------------------------------------------------------- backlinks


BACKLINK_OVERVIEW_FIELDS = [
    ("authority_score", "Authority score"),
    ("backlinks_total", "Backlinks (total)"),
    ("referring_domains", "Referring domains"),
    ("referring_ips", "Referring IPs"),
    ("follow_links", "Follow links"),
    ("nofollow_links", "Nofollow links"),
    ("organic_keywords", "Organic keywords"),
    ("organic_traffic", "Organic traffic (monthly)"),
    ("organic_cost", "Organic traffic value"),
    ("adwords_keywords", "Paid keywords"),
    ("rank", "Provider rank"),
]


def _authority_distribution(domain_rows: list[list[Any]]) -> list[list[Any]]:
    if not domain_rows:
        return []
    bands = [("0-19", 0, 19), ("20-39", 20, 39), ("40-59", 40, 59),
             ("60-79", 60, 79), ("80-100", 80, 100)]
    total = len(domain_rows)
    rows: list[list[Any]] = []
    for label, low, high in bands:
        matched = [
            row for row in domain_rows
            if isinstance(row[1], int | float) and low <= row[1] <= high
        ]
        backlinks = sum(row[2] for row in matched if isinstance(row[2], int | float))
        rows.append([
            label, len(matched), round(len(matched) / total, 3) if total else None,
            backlinks or 0,
        ])
    unscored = [row for row in domain_rows if not isinstance(row[1], int | float)]
    if unscored:
        rows.append(["Authority unavailable", len(unscored),
                     round(len(unscored) / total, 3), 0])
    return rows


def _link_gap_rows(data: dict[str, Any], available: bool) -> list[list[Any]]:
    if not available:
        return []
    return [
        [
            competitor.get("domain"),
            competitor.get("referring_domains"),
            competitor.get("common_keywords"),
            competitor.get("gap_keywords"),
            "Referring domains linking to this competitor are candidate outreach "
            "targets; each must be manually qualified before contact.",
            competitor.get("evidence_ids"),
        ]
        for competitor in _competitors(data)
    ]


def _backlink_competitor_rows(
    data: dict[str, Any], overview: dict[str, Any], available: bool
) -> list[list[Any]]:
    if not available:
        return []
    rows: list[list[Any]] = [[
        f"{data.get('client', {}).get('domain')} (client)",
        overview.get("authority_score"), overview.get("backlinks_total"),
        overview.get("referring_domains"), overview.get("organic_keywords"),
        overview.get("organic_traffic"),
    ]]
    for competitor in _competitors(data):
        rows.append([
            competitor.get("domain"), competitor.get("authority_score"),
            competitor.get("backlinks_total"), competitor.get("referring_domains"),
            competitor.get("organic_keywords"), competitor.get("organic_traffic"),
        ])
    return rows


def _backlink_audit(data: dict[str, Any], path: Path) -> Path:
    meta = _meta(data, "Backlink Audit Report")
    backlinks = _backlinks(data)
    available = _backlinks_available(data)
    reason = _unavailable_reason(backlinks, "no backlink provider connected")
    unavailable = f"{UNAVAILABLE} — {reason}"
    workbook = _new_workbook()

    overview = backlinks.get("overview") if isinstance(backlinks.get("overview"), dict) else {}
    _add_sheet(
        workbook, meta, "Overview",
        ["Metric", "Value", "Source"],
        [
            [label, overview.get(key), "Provider response"]
            for key, label in BACKLINK_OVERVIEW_FIELDS
        ] if available else [],
        [34, W_NUM + 8, W_TEXT],
        wrap_columns=frozenset({3}),
        empty_message=unavailable,
    )

    domain_rows = [
        [
            row.get("domain"), row.get("authority_score"), row.get("backlinks"),
            row.get("country"), row.get("first_seen"), row.get("last_seen"),
        ]
        for row in (backlinks.get("referring_domains") or [])
        if isinstance(row, dict)
    ]
    _add_sheet(
        workbook, meta, "Referring Domains",
        ["Referring Domain", "Authority Score", "Backlinks", "Country",
         "First Seen", "Last Seen"],
        domain_rows, [W_TEXT, W_NUM, W_NUM, W_NUM, W_DATE, W_DATE],
        empty_message=unavailable,
    )
    _add_sheet(
        workbook, meta, "Authority Distribution",
        ["Authority Band", "Referring Domains", "Share", "Backlinks"],
        _authority_distribution(domain_rows), [W_LABEL, W_NUM, W_NUM, W_NUM],
        empty_message=unavailable,
    )
    _add_sheet(
        workbook, meta, "Link Gap Opportunities",
        ["Competitor", "Referring Domains", "Common Keywords", "Gap Keywords",
         "Opportunity", "Evidence IDs"],
        _link_gap_rows(data, available), [W_TEXT, W_NUM, W_NUM, W_NUM, W_TEXT, 22],
        wrap_columns=frozenset({5}),
        empty_message=(
            unavailable if not available
            else "No competitor link gap was measurable from the connected provider."
        ),
    )
    _add_sheet(
        workbook, meta, "Competitor Comparison",
        ["Domain", "Authority Score", "Backlinks", "Referring Domains",
         "Organic Keywords", "Organic Traffic"],
        _backlink_competitor_rows(data, overview, available),
        [W_TEXT, W_NUM, W_NUM, W_NUM, W_NUM, W_NUM],
        empty_message=unavailable,
    )
    _add_methodology(
        workbook, meta, data,
        (
            "Referring-domain and authority metrics exactly as returned by the connected "
            "backlink provider. No domain is classified as toxic and no disavow list is "
            "produced without manual review."
            if available
            else "Backlink evidence was not collected for this run."
        ),
        extra=[
            ["Backlink status", backlinks.get("status") or UNAVAILABLE],
            ["Unavailable reason", reason if not available else "Not applicable"],
            [
                "What a connected provider adds",
                "A referring-domain inventory with authority scores and first/last seen "
                "dates, authority distribution, competitor link gaps and an anchor-text "
                "profile. Until a provider is approved these sheets stay empty rather "
                "than estimated.",
            ],
        ],
    )
    return _save(workbook, path)


# --------------------------------------------------------------------------- competitors


def _competitor_landscape(data: dict[str, Any], path: Path) -> Path:
    meta = _meta(data, "Competitor Landscape Analysis")
    competitors = _competitors(data)
    market = _market(data)
    domain_metrics = market.get("domain") if isinstance(market.get("domain"), dict) else {}
    available = bool(competitors)
    reason = _unavailable_reason(market, "no market data provider connected")
    unavailable = f"{UNAVAILABLE} — {reason}"
    workbook = _new_workbook()

    comparison: list[list[Any]] = []
    if available:
        comparison.append([
            f"{data.get('client', {}).get('domain')} (client)", "Client",
            domain_metrics.get("organic_keywords"), domain_metrics.get("organic_traffic"),
            domain_metrics.get("organic_cost"), domain_metrics.get("adwords_keywords"),
            domain_metrics.get("authority_score"), None, None,
        ])
        for competitor in competitors:
            comparison.append([
                competitor.get("domain"), "Competitor",
                competitor.get("organic_keywords"), competitor.get("organic_traffic"),
                competitor.get("organic_cost"), competitor.get("adwords_keywords"),
                competitor.get("authority_score"), competitor.get("relevance"),
                competitor.get("common_keywords"),
            ])
    _add_sheet(
        workbook, meta, "Domain Comparison",
        ["Domain", "Role", "Organic Keywords", "Organic Traffic", "Organic Cost",
         "Paid Keywords", "Authority Score", "Relevance", "Common Keywords"],
        comparison,
        [W_TEXT, W_LABEL, W_NUM, W_NUM, W_NUM, W_NUM, W_NUM, W_NUM, W_NUM],
        empty_message=unavailable,
    )
    _add_sheet(
        workbook, meta, "Shared Keyword Overlap",
        ["Competitor ID", "Domain", "Common Keywords", "Organic Keywords",
         "Relevance", "Evidence IDs"],
        [
            [
                competitor.get("id"), competitor.get("domain"),
                competitor.get("common_keywords"), competitor.get("organic_keywords"),
                competitor.get("relevance"), competitor.get("evidence_ids"),
            ]
            for competitor in competitors
        ],
        [W_NUM, W_TEXT, W_NUM, W_NUM, W_NUM, 22],
        empty_message=unavailable,
    )
    _add_sheet(
        workbook, meta, "Gap Opportunities",
        ["Competitor", "Gap Keywords", "Organic Traffic", "Organic Cost",
         "Interpretation", "Evidence IDs"],
        [
            [
                competitor.get("domain"), competitor.get("gap_keywords"),
                competitor.get("organic_traffic"), competitor.get("organic_cost"),
                "Keywords the competitor ranks for and the client does not, as reported "
                "by the provider. Each requires a page decision before it becomes work.",
                competitor.get("evidence_ids"),
            ]
            for competitor in competitors
            if competitor.get("gap_keywords") is not None
        ],
        [W_TEXT, W_NUM, W_NUM, W_NUM, W_TEXT, 22],
        wrap_columns=frozenset({5}),
        empty_message=(
            unavailable if not available
            else "The provider returned no keyword gap counts for these competitors."
        ),
    )

    performance = data.get("performance_vs_competitors")
    performance = performance if isinstance(performance, dict) else {}
    _add_sheet(
        workbook, meta, "Performance vs Competitors",
        ["Metric", "Client", "Competitor Median", "Best Competitor", "Best Value",
         "Position", "Note"],
        [
            [
                row.get("metric"), row.get("client"), row.get("competitor_median"),
                row.get("best_competitor"), row.get("best_value"), row.get("position"),
                row.get("note"),
            ]
            for row in (performance.get("metrics") or [])
            if isinstance(row, dict)
        ],
        [30, W_NUM, W_NUM + 6, W_TEXT, W_NUM, W_LABEL, W_TEXT],
        wrap_columns=frozenset({7}),
        empty_message=_unavailable_message(
            performance, "competitor performance was not measured for this run"
        ),
    )
    _add_methodology(
        workbook, meta, data,
        (
            "Competitor domains and their organic metrics exactly as returned by the "
            "connected market-data provider for the configured database. Relevance and "
            "common-keyword counts are provider values, not estimates."
            if available
            else "Competitor evidence was not collected for this run."
        ),
        extra=[
            ["Market status", market.get("status") or UNAVAILABLE],
            ["Market provider", market.get("provider") or UNAVAILABLE],
            ["Market database", market.get("database") or UNAVAILABLE],
            ["Provider units spent", market.get("units_spent")],
            ["Unavailable reason", reason if not available else "Not applicable"],
        ],
    )
    return _save(workbook, path)


# --------------------------------------------------------------------------- ecommerce


def _coverage_verdict(pages: list[dict[str, Any]], predicate: Any) -> str:
    if not pages:
        return f"{UNAVAILABLE} — no pages of this type were crawled"
    passing = len([page for page in pages if predicate(page)])
    return f"{passing}/{len(pages)} pages pass"


def _uniqueness_verdict(pages: list[dict[str, Any]], key: str) -> str:
    if not pages:
        return f"{UNAVAILABLE} — no pages of this type were crawled"
    populated = [
        value for value in
        (str(page.get(key) or "").strip().casefold() for page in pages)
        if value
    ]
    return f"{len(set(populated))} unique of {len(pages)} pages"


def _ecommerce_audit(data: dict[str, Any], path: Path) -> Path:
    meta = _meta(data, "Ecommerce Audit Report")
    pages = _pages(data)
    workbook = _new_workbook()

    products = [page for page in pages if str(page.get("page_type")) == "Product"]
    collections = [page for page in pages if str(page.get("page_type")) == "Collection"]

    _add_sheet(
        workbook, meta, "Product Pages",
        ["URL", "Title", "Title Length", "Meta Description", "H1", "Word Count",
         "Product Schema", "Images", "Images Missing Alt", "Status", "Evidence ID"],
        [
            [
                page.get("normalized_url"), page.get("title"),
                _text_length(page.get("title")), page.get("meta_description"),
                page.get("h1"), page.get("word_count"),
                "Yes" if _has_schema(page, "product") else "No",
                page.get("images_total"), page.get("images_missing_alt"),
                page.get("status_code"), page.get("evidence_id"),
            ]
            for page in products
        ],
        [W_URL, W_TEXT, W_NUM, W_TEXT, W_TEXT, W_NUM, W_LABEL, W_NUM, W_NUM, W_NUM, W_NUM],
        wrap_columns=frozenset({2, 4, 5}),
        empty_message="No product-pattern URLs were captured during the crawl window.",
    )
    _add_sheet(
        workbook, meta, "Collection Pages",
        ["URL", "Title", "Meta Description", "H1", "Word Count", "Internal Links",
         "Indexability", "Status", "Evidence ID"],
        [
            [
                page.get("normalized_url"), page.get("title"),
                page.get("meta_description"), page.get("h1"), page.get("word_count"),
                page.get("internal_links"), page.get("indexability"),
                page.get("status_code"), page.get("evidence_id"),
            ]
            for page in collections
        ],
        [W_URL, W_TEXT, W_TEXT, W_TEXT, W_NUM, W_NUM, 28, W_NUM, W_NUM],
        wrap_columns=frozenset({2, 3, 4}),
        empty_message="No collection-pattern URLs were captured during the crawl window.",
    )

    schema_rows: list[list[Any]] = []
    for label, group, expected in (
        ("Product", products, "Product"),
        ("Collection", collections, "ItemList / CollectionPage"),
    ):
        for page in group:
            declared = _schema_text(page)
            matched = _has_schema(page, expected.split(" / ")[0].casefold())
            schema_rows.append([
                page.get("normalized_url"), label, declared,
                "Yes" if matched else "No", expected,
                "Present" if matched
                else f"Add {expected} structured data to this page",
                page.get("evidence_id"),
            ])
    _add_sheet(
        workbook, meta, "Product Schema Coverage",
        ["URL", "Page Group", "Declared Schema Types", "Expected Schema Present",
         "Expected Schema", "Gap", "Evidence ID"],
        schema_rows, [W_URL, W_LABEL, 30, W_LABEL, 26, W_TEXT, W_NUM],
        wrap_columns=frozenset({3, 6}), color_column=4, color_map=ISSUE_COLORS,
        empty_message="No commerce pages were captured during the crawl window.",
    )
    _add_sheet(
        workbook, meta, "Checklist",
        ["Check", "Result", "Evidence Basis"],
        [
            ["Product schema present on product pages",
             _coverage_verdict(products, lambda page: _has_schema(page, "product")),
             "Crawl-observed JSON-LD / microdata types"],
            ["Unique titles on product pages",
             _uniqueness_verdict(products, "title"), "Crawl-observed <title> values"],
            ["Meta descriptions on product pages",
             _coverage_verdict(products, lambda page: bool(page.get("meta_description"))),
             "Crawl-observed meta description"],
            ["Single H1 on collection pages",
             _coverage_verdict(collections, lambda page: _h1_count(page) == 1),
             "Crawl-observed heading structure"],
            ["Collection pages indexable",
             _coverage_verdict(collections, lambda page: not _indexability_blocked(page)),
             "Crawl-observed robots directives"],
            ["Image alt coverage on product pages",
             _coverage_verdict(
                 products, lambda page: (page.get("images_missing_alt") or 0) == 0
             ),
             "Crawl-observed image inventory"],
            ["Checkout, cart and account funnel review",
             f"{UNAVAILABLE} — authenticated funnel pages are out of crawl scope",
             "Not measured"],
        ],
        [46, 34, W_TEXT], wrap_columns=frozenset({1, 2, 3}),
    )
    _add_methodology(
        workbook, meta, data,
        "Commerce page groups classified from URL patterns, then assessed against "
        "crawl-observed structured data, metadata and image inventories. Pricing, stock "
        "and transactional behaviour are not simulated.",
    )
    return _save(workbook, path)


# --------------------------------------------------------------------------- GEO / AEO


def _geo_recommendations(rows: list[list[Any]]) -> list[list[Any]]:
    if not rows:
        return []
    recommendations: list[list[Any]] = []
    no_schema = len([row for row in rows if row[2] == "No"])
    if no_schema:
        recommendations.append([
            "Add page-appropriate structured data (Organization, Product, Article or "
            "LocalBusiness) to pages with no detected schema.",
            no_schema, "Crawl detected no JSON-LD or microdata types on these pages.",
            "agency_admin",
        ])
    no_answer = len([row for row in rows if row[3] == "No"])
    if no_answer:
        recommendations.append([
            "Introduce FAQ or QA blocks with matching FAQPage/QAPage schema on pages "
            "that answer recurring buyer questions.",
            no_answer, "No answer-oriented schema type was detected in the crawl.",
            "editorial",
        ])
    bad_h1 = len([row for row in rows if row[4] != 1])
    if bad_h1:
        recommendations.append([
            "Give every page exactly one descriptive H1 so extractive answers have a "
            "single anchor heading.",
            bad_h1, "Crawl captured zero or multiple H1 elements on these pages.",
            "editorial",
        ])
    thin = len([row for row in rows if isinstance(row[5], int) and row[5] < 300])
    if thin:
        recommendations.append([
            "Expand short pages to at least 300 substantive words before expecting "
            "answer-engine citation.",
            thin, "Crawl-observed body word count below 300.", "editorial",
        ])
    no_meta = len([row for row in rows if row[6] == "No"])
    if no_meta:
        recommendations.append([
            "Write a meta description for every indexable page.",
            no_meta, "Crawl found no meta description element.", "editorial",
        ])
    return recommendations


def _geo_aeo(data: dict[str, Any], path: Path) -> Path:
    meta = _meta(data, "GEO & AEO Readiness Scorecard")
    pages = _pages(data)
    workbook = _new_workbook()

    rows: list[list[Any]] = []
    for page in pages:
        score, basis = _answer_ready(page)
        schema_types = {str(item).casefold() for item in (page.get("schema_types") or [])}
        rows.append([
            page.get("normalized_url"), page.get("page_type"),
            "Yes" if schema_types else "No",
            "Yes" if schema_types & ANSWER_SCHEMA else "No",
            _h1_count(page), page.get("word_count"),
            "Yes" if page.get("meta_description") else "No",
            score, basis, page.get("evidence_id"),
        ])
    _add_sheet(
        workbook, meta, "Page-Level Readiness",
        ["URL", "Page Type", "Structured Data", "Answer Schema", "H1 Count",
         "Word Count", "Meta Description", "Answer-Ready Score", "Score Basis",
         "Evidence ID"],
        rows,
        [W_URL, W_LABEL, W_LABEL, W_LABEL, W_NUM, W_NUM, W_LABEL, W_NUM, W_TEXT, W_NUM],
        wrap_columns=frozenset({9}),
        empty_message="No pages were captured during the crawl window.",
        note=(
            "Answer-ready score is a deterministic rule over crawl facts: structured data "
            "+30, answer schema +20, single H1 +20, 300+ words +15, meta description +15."
        ),
    )

    summary: list[list[Any]] = []
    if rows:
        scores = [row[7] for row in rows]
        for label, low, high in (
            ("80-100 Ready", 80, 100), ("60-79 Nearly ready", 60, 79),
            ("40-59 Partial", 40, 59), ("0-39 Not ready", 0, 39),
        ):
            matched = [score for score in scores if low <= score <= high]
            summary.append([label, len(matched), round(len(matched) / len(scores), 3)])
        summary.append(["Median score", sorted(scores)[len(scores) // 2], None])
        with_answer = len([row for row in rows if row[3] == "Yes"])
        summary.append([
            "Pages with answer schema", with_answer, round(with_answer / len(rows), 3),
        ])
    _add_sheet(
        workbook, meta, "Readiness Summary",
        ["Band", "Pages", "Share"], summary, [28, W_NUM, W_NUM],
        empty_message="No pages were captured during the crawl window.",
    )
    _add_sheet(
        workbook, meta, "Recommendations",
        ["Recommendation", "Pages Affected", "Basis", "Approval Class"],
        _geo_recommendations(rows), [W_TEXT, W_NUM, W_TEXT, 24],
        wrap_columns=frozenset({1, 3}),
        empty_message="Every crawled page already meets the answer-readiness rule.",
    )
    _add_methodology(
        workbook, meta, data,
        "Answer-engine readiness derived only from crawl facts: structured data types, "
        "answer-oriented schema, heading structure, body length and meta description. "
        "No third-party AI-visibility score is estimated.",
    )
    return _save(workbook, path)


# --------------------------------------------------------------------------- CRO / UX


def _quick_win_action(signal: str) -> str:
    actions = {
        "Missing H1": "Add one descriptive H1 that states the page's offer or answer.",
        "Thin money page": "Expand the page with specification, proof and FAQ content.",
        "Slow response": "Investigate server response time before layout optimisation.",
        "Images without alt text": "Write descriptive alt text for every meaningful image.",
        "No meta description": "Write a unique meta description within the length budget.",
        "No viewport meta tag": "Add a responsive viewport meta tag to the document head.",
    }
    return actions.get(signal, "Review with the delivery team before scheduling.")


def _cro_findings(pages: list[dict[str, Any]]) -> list[list[Any]]:
    findings: list[list[Any]] = []
    for page in pages:
        page_type = str(page.get("page_type") or "")
        money = page_type in MONEY_PAGE_TYPES
        url = page.get("normalized_url")
        evidence = page.get("evidence_id")
        if _h1_count(page) == 0:
            findings.append([
                url, page_type, "Missing H1", "High" if money else "Medium",
                "No H1 element was captured, so the page offers no primary heading cue.",
                evidence,
            ])
        word_count = page.get("word_count")
        if money and isinstance(word_count, int):
            # 150-250 words is NORMAL for ecommerce product templates — only
            # genuinely thin pages are defects, and only near-empty ones are
            # severe. This intentionally diverges from the old blanket
            # THIN_WORDS<250 High, which drowned the sheet in noise.
            threshold = (
                THIN_WORDS_CRITICAL
                if page_type in {"Product", "Collection"}
                else THIN_WORDS_CONTEXT
            )
            if word_count < threshold:
                severity = "High" if word_count < 60 or _h1_count(page) == 0 else "Medium"
                findings.append([
                    url, page_type, "Thin money page", severity,
                    f"{word_count} observed body words on a revenue-critical page.",
                    evidence,
                ])
        response = page.get("response_ms")
        if isinstance(response, int | float) and response >= SLOW_MS:
            findings.append([
                url, page_type, "Slow response", "High" if money else "Medium",
                f"Server responded in {response} ms during the crawl.", evidence,
            ])
        missing_alt = page.get("images_missing_alt")
        if isinstance(missing_alt, int) and missing_alt > 0:
            findings.append([
                url, page_type, "Images without alt text", "Medium",
                f"{missing_alt} of {page.get('images_total')} images had no alt text.",
                evidence,
            ])
        if not page.get("meta_description"):
            findings.append([
                url, page_type, "No meta description", "Low",
                "No meta description was captured, weakening the search snippet.",
                evidence,
            ])
        if _facts(page).get("has_viewport") is False:
            findings.append([
                url, page_type, "No viewport meta tag", "High",
                "No responsive viewport meta tag was captured on this page.", evidence,
            ])
    return findings


def _aggregate_signal_floods(findings: list[list[Any]]) -> list[list[Any]]:
    """Collapse a signal that fires across many pages into one summary row.

    One systemic template issue must read as ONE finding with a page count,
    not a wall of identical rows — reviewers stop reading after the third.
    """

    by_signal: dict[str, list[list[Any]]] = {}
    for row in findings:
        by_signal.setdefault(str(row[2]), []).append(row)
    out: list[list[Any]] = []
    for signal, rows in by_signal.items():
        if len(rows) <= SIGNAL_AGGREGATE_THRESHOLD:
            out.extend(rows)
            continue
        worst = "High" if any(row[3] == "High" for row in rows) else rows[0][3]
        sample = ", ".join(str(row[0]) for row in rows[:3])
        page_types = sorted({str(row[1]) for row in rows if row[1]})
        out.append([
            f"{len(rows)} pages (systemic)",
            "/".join(page_types)[:40] or "Various",
            signal,
            worst,
            f"{len(rows)} pages share this signal — a template-level fix, "
            f"not {len(rows)} separate tasks. Examples: {sample}. "
            "Full page list in the Technical and Content audit workbooks.",
            rows[0][5],
        ])
    return out


def _cro_ux(data: dict[str, Any], path: Path) -> Path:
    meta = _meta(data, "CRO & UX Findings")
    workbook = _new_workbook()
    raw_findings = _cro_findings(_pages(data))
    findings = _aggregate_signal_floods(raw_findings)

    _add_sheet(
        workbook, meta, "Findings",
        ["URL", "Page Type", "Signal", "Severity", "Observation", "Evidence ID"],
        findings, [W_URL, W_LABEL, 26, W_NUM, W_TEXT, W_NUM],
        wrap_columns=frozenset({5}), color_column=4, color_map=SEVERITY_COLORS,
        empty_message="No crawl-derivable conversion or usability defects were observed.",
    )
    signal_counts = Counter(row[2] for row in raw_findings)
    _add_sheet(
        workbook, meta, "Quick Wins",
        ["Signal", "Pages Affected", "Impact", "Recommended Action", "Basis"],
        [
            [
                signal, count,
                "High" if signal in {"Missing H1", "Slow response"} else "Medium",
                _quick_win_action(signal), "Crawl-observed",
            ]
            for signal, count in sorted(signal_counts.items(), key=lambda item: -item[1])
        ],
        [26, W_NUM, W_NUM, W_TEXT, W_LABEL],
        wrap_columns=frozenset({4}), color_column=3, color_map=SEVERITY_COLORS,
        empty_message="No crawl-derivable quick wins were identified.",
    )
    _add_methodology(
        workbook, meta, data,
        "Conversion and usability signals derivable from crawl evidence only: heading "
        "structure, body depth on revenue pages, server response time, image alt "
        "coverage, snippet metadata and viewport declaration. Session behaviour, "
        "heatmaps and funnel analytics are not simulated.",
    )
    return _save(workbook, path)


# --------------------------------------------------------------------------- tracking


def _tracking_audit(data: dict[str, Any], path: Path) -> Path:
    meta = _meta(data, "Tracking Audit Report")
    pages = _pages(data)
    workbook = _new_workbook()
    tagged = [page for page in pages if "analytics_tags" in page]
    missing = f"{UNAVAILABLE} — per-page analytics detection was not captured for this run."

    _add_sheet(
        workbook, meta, "Tag Coverage",
        ["URL", "Page Type", "Analytics Tags", "Detected", "Tag Count", "Status",
         "Evidence ID"],
        [
            [
                page.get("normalized_url"), page.get("page_type"),
                ", ".join(page.get("analytics_tags") or []) or "No tags detected",
                "Yes" if page.get("analytics_tags") else "No",
                len(page.get("analytics_tags") or []),
                page.get("status_code"), page.get("evidence_id"),
            ]
            for page in tagged
        ],
        [W_URL, W_LABEL, W_TEXT, W_LABEL, W_NUM, W_NUM, W_NUM],
        wrap_columns=frozenset({3}), empty_message=missing,
    )

    summary: list[list[Any]] = []
    if tagged:
        counter: Counter[str] = Counter()
        for page in tagged:
            for tag in page.get("analytics_tags") or []:
                counter[str(tag)] += 1
        detected = len([page for page in tagged if page.get("analytics_tags")])
        summary.append(["Pages crawled with tag detection", len(tagged), None])
        summary.append(
            ["Pages with at least one tag", detected, round(detected / len(tagged), 3)]
        )
        summary.append([
            "Pages with no tag detected", len(tagged) - detected,
            round((len(tagged) - detected) / len(tagged), 3),
        ])
        for tag, count in sorted(counter.items(), key=lambda item: (-item[1], item[0])):
            summary.append([f"Tag: {tag}", count, round(count / len(tagged), 3)])
        summary.append([
            "Server-side and consent-mode configuration",
            f"{UNAVAILABLE} — not observable from HTML alone", None,
        ])
    _add_sheet(
        workbook, meta, "Summary",
        ["Measure", "Value", "Share"], summary, [46, W_NUM, W_NUM],
        empty_message=missing,
    )
    _add_sheet(
        workbook, meta, "Tracking Findings",
        FINDING_HEADERS, _finding_rows(_findings_for(data, "analytics", "tracking")),
        FINDING_WIDTHS, wrap_columns=frozenset({4, 5, 6}),
        color_column=3, color_map=SEVERITY_COLORS,
        empty_message="No analytics or tracking findings were raised for this run.",
    )
    _add_methodology(
        workbook, meta, data,
        "Analytics and tag-manager signatures detected in the crawled HTML of each page. "
        "Container configuration, event schemas and consent mode require platform access "
        "and are reported as unavailable rather than assumed.",
    )
    return _save(workbook, path)


# --------------------------------------------------------------------------- GBP local


def _gbp_local(data: dict[str, Any], path: Path) -> Path:
    meta = _meta(data, "GBP & Local Audit")
    pages = _pages(data)
    workbook = _new_workbook()

    rows: list[list[Any]] = []
    for page in pages:
        schema_types = {str(item).casefold() for item in (page.get("schema_types") or [])}
        local_schema = sorted(schema_types & LOCAL_SCHEMA)
        contact_page = "/contact" in str(page.get("normalized_url") or "").casefold()
        if not local_schema and not contact_page:
            continue
        rows.append([
            page.get("normalized_url"), page.get("page_type"),
            ", ".join(local_schema) or "None detected",
            "Yes" if contact_page else "No",
            "Yes" if "postaladdress" in schema_types
            else f"{UNAVAILABLE} — no address schema",
            page.get("word_count"), page.get("evidence_id"),
        ])
    _add_sheet(
        workbook, meta, "Local Signals",
        ["URL", "Page Type", "Local Schema Types", "Contact Page", "Address Schema",
         "Word Count", "Evidence ID"],
        rows, [W_URL, W_LABEL, 30, W_LABEL, 30, W_NUM, W_NUM],
        empty_message="No local business schema or contact pages were observed in the crawl.",
    )

    has_org = any(
        "organization" in str(item).casefold()
        for page in pages for item in (page.get("schema_types") or [])
    )
    has_address = any(
        "postaladdress" in str(item).casefold()
        for page in pages for item in (page.get("schema_types") or [])
    )
    _add_sheet(
        workbook, meta, "NAP Observations",
        ["Signal", "Observation", "Evidence Basis"],
        [
            ["Business name in structured data", "Yes" if has_org else "No",
             "Crawl-observed schema types"],
            ["Postal address in structured data", "Yes" if has_address else "No",
             "Crawl-observed schema types"],
            ["Telephone in structured data",
             f"{UNAVAILABLE} — telephone properties are not captured by the crawler",
             "Not measured"],
            ["Google Business Profile listing",
             f"{UNAVAILABLE} — GBP API access has not been approved for this run",
             "Not measured"],
            ["Review and rating signals",
             f"{UNAVAILABLE} — review platforms are outside the approved crawl boundary",
             "Not measured"],
        ],
        [40, 46, 34], wrap_columns=frozenset({1, 2, 3}),
    )
    _add_sheet(
        workbook, meta, "Recommendations",
        ["Recommendation", "Approval Class", "Basis"],
        [
            ["Publish LocalBusiness structured data with the full NAP block on the "
             "contact and homepage.", "agency_admin",
             "No LocalBusiness schema was detected." if not has_org
             else "Organization schema is present; extend it to LocalBusiness."],
            ["Keep the on-page NAP string byte-identical to the Google Business Profile "
             "listing once GBP access is approved.", "agency_admin",
             "GBP data is unavailable for this run."],
            ["Create or verify a location page per serviced area before expanding local "
             "content.", "editorial", "Crawl-observed information architecture."],
        ],
        [W_TEXT + 15, 24, W_TEXT], wrap_columns=frozenset({1, 3}),
    )
    _add_methodology(
        workbook, meta, data,
        "Local signals observed in the approved-domain crawl: local structured-data "
        "types, contact-page presence and address schema. Google Business Profile "
        "content, reviews and citation listings require external access and are "
        "reported as unavailable.",
    )
    return _save(workbook, path)


# --------------------------------------------------------------------------- baseline


def _baseline_performance(data: dict[str, Any], path: Path) -> Path:
    meta = _meta(data, "Baseline Performance Analysis")
    pages = _pages(data)
    market = _market(data)
    workbook = _new_workbook()

    timed = [page for page in pages if isinstance(page.get("response_ms"), int | float)]
    _add_sheet(
        workbook, meta, "Response Times",
        ["URL", "Response (ms)", "Status Code", "Page Type", "Evidence ID"],
        [
            [page.get("normalized_url"), page.get("response_ms"), page.get("status_code"),
             page.get("page_type"), page.get("evidence_id")]
            for page in sorted(timed, key=lambda item: item.get("response_ms") or 0,
                               reverse=True)
        ],
        [W_URL, W_NUM, W_NUM, W_LABEL, W_NUM],
        empty_message=f"{UNAVAILABLE} — response timing was not captured for this run.",
    )

    weighed = [page for page in pages if isinstance(page.get("body_bytes"), int | float)]
    _add_sheet(
        workbook, meta, "Page Weight",
        ["URL", "Body Bytes", "Body KB", "Page Type", "Evidence ID"],
        [
            [page.get("normalized_url"), page.get("body_bytes"),
             round((page.get("body_bytes") or 0) / 1024, 1), page.get("page_type"),
             page.get("evidence_id")]
            for page in sorted(weighed, key=lambda item: item.get("body_bytes") or 0,
                               reverse=True)
        ],
        [W_URL, W_NUM, W_NUM, W_LABEL, W_NUM],
        empty_message=f"{UNAVAILABLE} — HTML payload size was not captured for this run.",
    )

    distribution: list[list[Any]] = []
    if timed:
        values = sorted(float(page["response_ms"]) for page in timed)
        for label, low, high in (
            ("Under 200 ms", 0, 199), ("200-499 ms", 200, 499),
            ("500-999 ms", 500, 999), ("1000-1499 ms", 1000, 1499),
            ("1500 ms and above", 1500, 10**9),
        ):
            matched = [value for value in values if low <= value <= high]
            distribution.append([label, len(matched), round(len(matched) / len(values), 3)])
        distribution.append(["Median response (ms)", values[len(values) // 2], None])
        distribution.append(["Slowest response (ms)", values[-1], None])
    if weighed:
        kb = sorted(float(page["body_bytes"]) / 1024 for page in weighed)
        distribution.append(["Median HTML weight (KB)", round(kb[len(kb) // 2], 1), None])
        distribution.append(["Heaviest HTML (KB)", round(kb[-1], 1), None])
    _add_sheet(
        workbook, meta, "Distribution",
        ["Measure", "Value", "Share"], distribution, [34, W_NUM, W_NUM],
        empty_message=f"{UNAVAILABLE} — no timing or weight samples were captured.",
    )

    domain_metrics = market.get("domain") if isinstance(market.get("domain"), dict) else {}
    _add_sheet(
        workbook, meta, "Market Baseline",
        ["Metric", "Value", "Provider", "Fetched At"],
        [
            [label, domain_metrics.get(key), market.get("provider") or UNAVAILABLE,
             market.get("fetched_at") or UNAVAILABLE]
            for key, label in BACKLINK_OVERVIEW_FIELDS
        ] if str(market.get("status") or "").casefold() == "available" else [],
        [34, W_NUM + 8, W_LABEL, 24],
        empty_message=_unavailable_message(market, "no market data provider connected"),
    )
    _add_methodology(
        workbook, meta, data,
        "Server response timing and HTML payload weight as observed by the crawler, plus "
        "the connected provider's domain baseline when one is approved. Lab and field "
        "Core Web Vitals require a PageSpeed or CrUX connection and are not simulated.",
    )
    return _save(workbook, path)


# --------------------------------------------------------------------------- keywords


def _master_keyword_universe(data: dict[str, Any], path: Path) -> Path:
    meta = _meta(data, "Master Keyword Universe")
    keywords = _keywords(data)
    clusters = _clusters(data)
    market = _market(data)
    workbook = _new_workbook()
    unavailable = _unavailable_message(market, "no keyword provider connected")

    volumes = [k.get("search_volume") for k in keywords
               if isinstance(k.get("search_volume"), int | float)]
    ranked = [k for k in keywords if isinstance(k.get("position"), int)]
    mapped = [k for k in keywords if str(k.get("landing_url") or "").strip()]
    summary: list[list[Any]] = []
    if keywords:
        summary = [
            ["Keywords in universe", len(keywords), "Provider response"],
            ["Keywords with a measured volume", len(volumes),
             "Only provider-supplied volumes are counted"],
            ["Total measured monthly volume", sum(volumes) if volumes else None,
             "Sum of provider volumes; blank volumes are excluded"],
            ["Keywords with a measured position", len(ranked), "Provider response"],
            ["Keywords mapped to a crawled URL", len(mapped),
             "Provider landing URL matched against the crawl"],
            ["Keywords with no mapped URL", len(keywords) - len(mapped),
             "These are the content gap"],
            ["Clusters identified", len(clusters), "Deterministic clustering rule"],
            ["Provider", market.get("provider") or UNAVAILABLE, "Market source"],
            ["Database", market.get("database") or UNAVAILABLE, "Market source"],
            ["Fetched at", market.get("fetched_at") or UNAVAILABLE, "Market source"],
        ]
    _add_sheet(
        workbook, meta, "Executive Summary",
        ["Measure", "Value", "Basis"], summary, [40, W_NUM + 8, W_TEXT],
        wrap_columns=frozenset({3}), empty_message=unavailable,
    )

    _add_sheet(
        workbook, meta, "Keyword Research Mapping",
        ["Keyword ID", "Phrase", "Position", "Previous Position", "Search Volume",
         "CPC", "Competition", "Results", "Traffic Share", "Traffic Cost Share",
         "Trend", "Landing URL", "Intent", "Funnel Stage", "Cluster", "Page Type",
         "Opportunity", "Source", "Evidence IDs"],
        [
            [
                keyword.get("id"), keyword.get("phrase"), keyword.get("position"),
                keyword.get("previous_position"), keyword.get("search_volume"),
                keyword.get("cpc"), keyword.get("competition"),
                keyword.get("results_count"), keyword.get("traffic_share"),
                keyword.get("traffic_cost_share"), keyword.get("trend"),
                keyword.get("landing_url"), keyword.get("intent"),
                keyword.get("funnel_stage"), keyword.get("cluster"),
                keyword.get("page_type"), keyword.get("opportunity"),
                keyword.get("source"), keyword.get("evidence_ids"),
            ]
            for keyword in keywords
        ],
        [W_NUM, W_TEXT, W_NUM, W_NUM, W_NUM, W_NUM, W_NUM, W_NUM, W_NUM, W_NUM,
         W_LABEL, W_URL, W_LABEL, W_LABEL, 24, W_LABEL, W_TEXT, W_LABEL, 22],
        wrap_columns=frozenset({2, 17}),
        empty_message=unavailable,
        note=None if keywords else KEYWORD_VOLUME_NOTE,
    )

    _add_sheet(
        workbook, meta, "Category & URL Mapping",
        ["Cluster ID", "Cluster", "Keywords", "Total Volume", "Primary URL",
         "Intent", "Coverage", "Evidence IDs"],
        [
            [
                cluster.get("id"), cluster.get("name"), cluster.get("keyword_count"),
                cluster.get("total_volume"), cluster.get("primary_url"),
                cluster.get("intent"), cluster.get("coverage"),
                cluster.get("evidence_ids"),
            ]
            for cluster in clusters
        ],
        [W_NUM, 30, W_NUM, W_NUM, W_URL, W_LABEL, W_LABEL, 22],
        color_column=7, color_map=COVERAGE_COLORS,
        empty_message=unavailable,
    )

    url_rows: dict[str, dict[str, Any]] = {}
    for keyword in mapped:
        url = str(keyword.get("landing_url"))
        entry = url_rows.setdefault(
            url, {"count": 0, "volume": 0, "phrases": [], "best": None}
        )
        entry["count"] += 1
        volume = keyword.get("search_volume")
        if isinstance(volume, int | float):
            entry["volume"] += volume
        entry["phrases"].append(str(keyword.get("phrase")))
        position = keyword.get("position")
        if isinstance(position, int) and (
            entry["best"] is None or position < entry["best"]
        ):
            entry["best"] = position
    crawled = {str(page.get("normalized_url")) for page in _pages(data)}
    _add_sheet(
        workbook, meta, "URL Mapping",
        ["Landing URL", "In Crawl", "Keywords", "Total Volume", "Best Position",
         "Sample Keywords"],
        [
            [
                url, "Yes" if url in crawled else "No", entry["count"],
                entry["volume"] or None, entry["best"],
                ", ".join(entry["phrases"][:8]),
            ]
            for url, entry in sorted(url_rows.items())
        ],
        [W_URL, W_LABEL, W_NUM, W_NUM, W_NUM, W_TEXT],
        wrap_columns=frozenset({6}), empty_message=unavailable,
    )

    funnel_rows: list[list[Any]] = []
    if keywords:
        stages = Counter(
            str(keyword.get("funnel_stage") or "Unclassified") for keyword in keywords
        )
        for stage in ("TOFU", "MOFU", "BOFU", "Unclassified"):
            count = stages.get(stage, 0)
            if not count and stage == "Unclassified":
                continue
            stage_keywords = [
                keyword for keyword in keywords
                if str(keyword.get("funnel_stage") or "Unclassified") == stage
            ]
            stage_volumes = [
                keyword.get("search_volume") for keyword in stage_keywords
                if isinstance(keyword.get("search_volume"), int | float)
            ]
            funnel_rows.append([
                stage, count, round(count / len(keywords), 3),
                sum(stage_volumes) if stage_volumes else None,
                len([
                    keyword for keyword in stage_keywords
                    if str(keyword.get("landing_url") or "").strip()
                ]),
            ])
    _add_sheet(
        workbook, meta, "Funnel Distribution",
        ["Funnel Stage", "Keywords", "Share", "Measured Volume", "Mapped to a URL"],
        funnel_rows, [W_LABEL, W_NUM, W_NUM, W_NUM, W_NUM],
        empty_message=unavailable,
    )
    _add_methodology(
        workbook, meta, data,
        "Every keyword row is a provider response or a crawl observation; positions, "
        "volumes, CPC and competition are never estimated. Funnel stage, cluster and "
        "opportunity are deterministic classifications over those measured values.",
        extra=[
            ["Market status", market.get("status") or UNAVAILABLE],
            ["Market provider", market.get("provider") or UNAVAILABLE],
            ["Provider units spent", market.get("units_spent")],
            ["Unavailable reason", market.get("unavailable_reason") or "Not applicable"],
        ],
    )
    return _save(workbook, path)


def _content_gap_analysis(data: dict[str, Any], path: Path) -> Path:
    meta = _meta(data, "Content Gap Analysis")
    keywords = _keywords(data)
    clusters = _clusters(data)
    market = _market(data)
    workbook = _new_workbook()
    unavailable = _unavailable_message(market, "no keyword provider connected")

    _add_sheet(
        workbook, meta, "Gap Keywords",
        ["Keyword ID", "Phrase", "Search Volume", "CPC", "Competition", "Intent",
         "Funnel Stage", "Cluster", "Recommended Page Type", "Opportunity",
         "Evidence IDs"],
        [
            [
                keyword.get("id"), keyword.get("phrase"), keyword.get("search_volume"),
                keyword.get("cpc"), keyword.get("competition"), keyword.get("intent"),
                keyword.get("funnel_stage"), keyword.get("cluster"),
                keyword.get("page_type"), keyword.get("opportunity"),
                keyword.get("evidence_ids"),
            ]
            for keyword in keywords
            if not str(keyword.get("landing_url") or "").strip()
        ],
        [W_NUM, W_TEXT, W_NUM, W_NUM, W_NUM, W_LABEL, W_LABEL, 24, W_LABEL, W_TEXT, 22],
        wrap_columns=frozenset({2, 10}),
        empty_message=(
            unavailable if not keywords
            else "Every measured keyword already maps to a crawled URL."
        ),
    )
    _add_sheet(
        workbook, meta, "Cluster Coverage",
        ["Cluster ID", "Cluster", "Coverage", "Keywords", "Total Volume",
         "Primary URL", "Intent", "Evidence IDs"],
        [
            [
                cluster.get("id"), cluster.get("name"), cluster.get("coverage"),
                cluster.get("keyword_count"), cluster.get("total_volume"),
                cluster.get("primary_url"), cluster.get("intent"),
                cluster.get("evidence_ids"),
            ]
            for cluster in clusters
        ],
        [W_NUM, 30, W_LABEL, W_NUM, W_NUM, W_URL, W_LABEL, 22],
        color_column=3, color_map=COVERAGE_COLORS,
        empty_message=unavailable,
    )
    _add_sheet(
        workbook, meta, "Priority Recommendations",
        ["Cluster", "Coverage", "Keywords", "Total Volume", "Recommended Action",
         "Target URL", "Evidence IDs"],
        [
            [
                cluster.get("name"), cluster.get("coverage"),
                cluster.get("keyword_count"), cluster.get("total_volume"),
                "Create a dedicated page"
                if str(cluster.get("coverage")).casefold() == "gap"
                else "Expand the existing page",
                cluster.get("primary_url") or "No page currently targets this cluster",
                cluster.get("evidence_ids"),
            ]
            for cluster in clusters
            if str(cluster.get("coverage") or "").casefold() != "covered"
        ],
        [30, W_LABEL, W_NUM, W_NUM, 30, W_URL, 22],
        empty_message=(
            unavailable if not clusters
            else "Every measured cluster is already covered by a crawled page."
        ),
    )
    _add_methodology(
        workbook, meta, data,
        "A gap is a measured keyword with no crawled landing URL, or a cluster whose "
        "coverage rule resolved to partial or gap. Nothing enters this list without a "
        "provider row behind it.",
    )
    return _save(workbook, path)


def _content_strategy(data: dict[str, Any], path: Path) -> Path:
    meta = _meta(data, "Content Strategy")
    clusters = _clusters(data)
    assets = list(data.get("content_assets") or [])
    workbook = _new_workbook()

    _add_sheet(
        workbook, meta, "Content Plan",
        ["Asset ID", "Title", "Asset Type", "Target URL", "Audience", "Intent",
         "Primary Topic", "Approval State", "Generation Method", "Evidence IDs"],
        [
            [
                asset.get("id"), asset.get("title"), asset.get("asset_type"),
                asset.get("target_url"), asset.get("audience"), asset.get("intent"),
                asset.get("primary_topic"), asset.get("approval_state"),
                asset.get("generation_method"), asset.get("evidence_ids"),
            ]
            for asset in assets
        ],
        [W_NUM, W_TEXT, W_LABEL, W_URL, W_LABEL, W_LABEL, 24, 26, W_LABEL, 22],
        wrap_columns=frozenset({2}),
        empty_message="No content assets cleared evidence checks during this run.",
    )
    _add_sheet(
        workbook, meta, "Cluster Roadmap",
        ["Cluster ID", "Cluster", "Intent", "Coverage", "Keywords", "Total Volume",
         "Primary URL", "Next Move"],
        [
            [
                cluster.get("id"), cluster.get("name"), cluster.get("intent"),
                cluster.get("coverage"), cluster.get("keyword_count"),
                cluster.get("total_volume"), cluster.get("primary_url"),
                {
                    "covered": "Maintain and refresh on the measurement cadence",
                    "partial": "Expand the existing page to cover the full cluster",
                    "gap": "Commission one new page for this cluster",
                }.get(str(cluster.get("coverage") or "").casefold(),
                      "Decide after the next evidence refresh"),
            ]
            for cluster in clusters
        ],
        [W_NUM, 30, W_LABEL, W_LABEL, W_NUM, W_NUM, W_URL, W_TEXT],
        wrap_columns=frozenset({8}),
        empty_message=_unavailable_message(_market(data), "no keyword provider connected"),
    )
    _add_sheet(
        workbook, meta, "Editorial Standards",
        ["Standard", "Status", "Reason"],
        [
            ["Every asset states its evidence basis", "Required",
             "Claims without an evidence ID are removed before release."],
            ["One primary intent per page", "Required",
             "Prevents the cannibalisation signals recorded in the content audit."],
            ["Single descriptive H1", "Required",
             "Matches the on-page rule applied across the technical audit."],
            ["Meta description within the length budget", "Required",
             f"{META_MIN}-{META_MAX} characters."],
            ["Internal link to the cluster's primary URL", "Required",
             "Keeps cluster authority consolidated on one target."],
            ["No performance forecast without a measured baseline", "Required",
             "Forecasts are withheld until first-party analytics are connected."],
        ],
        [46, W_LABEL, W_TEXT], wrap_columns=frozenset({1, 3}),
    )
    _add_methodology(
        workbook, meta, data,
        "Content planning derived from measured clusters and the evidence-checked asset "
        "register. Editorial standards restate the audit's own rules as production "
        "requirements.",
    )
    return _save(workbook, path)


def _url_architecture(data: dict[str, Any], path: Path) -> Path:
    meta = _meta(data, "URL Architecture Map")
    pages = _pages(data)
    workbook = _new_workbook()

    _add_sheet(
        workbook, meta, "URL Map",
        ["URL", "Page Type", "Depth", "Status", "Internal Links", "External Links",
         "Indexability", "Title"],
        [
            [
                page.get("normalized_url"), page.get("page_type"),
                _url_depth(str(page.get("normalized_url") or "")),
                page.get("status_code"), page.get("internal_links"),
                page.get("external_links"), page.get("indexability"), page.get("title"),
            ]
            for page in pages
        ],
        [W_URL, W_LABEL, W_NUM, W_NUM, W_NUM, W_NUM, 28, W_TEXT],
        wrap_columns=frozenset({8}),
        empty_message="No pages were captured during the crawl window.",
    )

    sections: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for page in pages:
        sections[str(page.get("page_type") or "Other")].append(page)
    rollup: list[list[Any]] = []
    for section in sorted(sections):
        group = sections[section]
        links = [p.get("internal_links") for p in group
                 if isinstance(p.get("internal_links"), int | float)]
        depths = [_url_depth(str(p.get("normalized_url") or "")) for p in group]
        words = [p.get("word_count") for p in group if isinstance(p.get("word_count"), int)]
        rollup.append([
            section, len(group),
            round(sum(links) / len(links), 1) if links else None,
            round(sum(depths) / len(depths), 1) if depths else None,
            round(sum(words) / len(words)) if words else None,
        ])
    _add_sheet(
        workbook, meta, "Sections",
        ["Page Type", "Pages", "Avg Internal Links", "Avg Depth", "Avg Word Count"],
        rollup, [W_LABEL + 8, W_NUM, W_NUM + 6, W_NUM, W_NUM + 4],
        empty_message="No pages were captured during the crawl window.",
    )

    depth_counter: Counter[int] = Counter(
        _url_depth(str(page.get("normalized_url") or "")) for page in pages
    )
    _add_sheet(
        workbook, meta, "Depth Distribution",
        ["Click Depth", "Pages", "Share"],
        [
            [depth, count, round(count / len(pages), 3)]
            for depth, count in sorted(depth_counter.items())
        ],
        [W_NUM, W_NUM, W_NUM],
        empty_message="No pages were captured during the crawl window.",
    )
    _add_methodology(
        workbook, meta, data,
        "Normalized URL inventory with path depth, observed status, link counts and "
        "indexability, rolled up by classified page type.",
    )
    return _save(workbook, path)


def _cannibalization_plan(data: dict[str, Any], path: Path) -> Path:
    meta = _meta(data, "Cannibalization Resolution Plan")
    workbook = _new_workbook()
    rows = _cannibalisation_rows(data)

    _add_sheet(
        workbook, meta, "Overlap Signals",
        ["Signal", "Shared Value", "Pages", "URLs", "Recommended Resolution",
         "Evidence IDs"],
        rows, [W_LABEL, W_TEXT, W_NUM, W_URL + 20, W_TEXT, 22],
        wrap_columns=frozenset({2, 4, 5}),
        empty_message="No overlapping targeting signals were observed in the crawl.",
    )
    _add_sheet(
        workbook, meta, "Resolution Plan",
        ["#", "Signal", "Shared Value", "Pages", "Decision Required", "Resolution",
         "Approval Class", "Evidence IDs"],
        [
            [index, row[0], row[1], row[2], "Choose one canonical target", row[4],
             "editorial", row[5]]
            for index, row in enumerate(rows, start=1)
        ],
        [6, W_LABEL, W_TEXT, W_NUM, 26, W_TEXT, 20, 22],
        wrap_columns=frozenset({3, 6}),
        empty_message="No resolution work is required from the observed evidence.",
    )
    _add_sheet(
        workbook, meta, "Cannibalisation Findings",
        FINDING_HEADERS,
        _finding_rows(
            _findings_for(data, "keyword_architecture", "cannibalization", "keywords")
        ),
        FINDING_WIDTHS, wrap_columns=frozenset({4, 5, 6}),
        color_column=3, color_map=SEVERITY_COLORS,
        empty_message="No cannibalisation findings were raised for this run.",
    )
    _add_methodology(
        workbook, meta, data,
        "Overlap is asserted only from observed duplicate titles or from one measured "
        "keyword mapping to more than one crawled URL. No overlap is inferred from topic "
        "similarity alone.",
    )
    return _save(workbook, path)


# --------------------------------------------------------------------------- plan


ACTION_PLAN_HEADERS = [
    "Phase", "Week", "Task #", "Category", "Description", "Pages/Items", "Priority",
    "Est. Effort", "Owner", "Deliverable", "KPI / Success Metric", "Approval Class", "Notes",
]


def _category_label(raw: str) -> str:
    return raw.replace("_", " ").replace("-", " ").strip().title() or "General"


def _action_plan(data: dict[str, Any], path: Path) -> Path:
    meta = _meta(data, "16-Week Action Plan")
    actions = list(data.get("actions") or [])
    workbook = _new_workbook()

    finding_by_evidence: dict[str, dict[str, Any]] = {}
    for finding in data.get("findings") or []:
        for evidence_id in finding.get("evidence_ids") or []:
            finding_by_evidence.setdefault(str(evidence_id), finding)

    plan_rows: list[list[Any]] = []
    for action in actions:
        matched: dict[str, Any] | None = None
        for evidence_id in action.get("evidence_ids") or []:
            matched = finding_by_evidence.get(str(evidence_id))
            if matched is not None:
                break
        explicit = str(action.get("category") or "").strip()
        category = explicit or (
            _category_label(str(matched.get("category") or "")) if matched else "General"
        )
        week = action.get("week")
        week_end = action.get("week_end") or week
        week_label = (
            f"W{week}" if week == week_end or week_end is None else f"W{week}-W{week_end}"
        )
        plan_rows.append([
            action.get("phase"),
            week_label if week is not None else None,
            action.get("id"),
            category,
            action.get("action"),
            (matched or {}).get("affected_count"),
            action.get("priority"),
            action.get("effort"),
            action.get("owner"),
            action.get("deliverable") or "Unavailable - not specified in the plan record",
            action.get("kpi"),
            action.get("approval_class"),
            action.get("notes"),
        ])
    _add_sheet(
        workbook, meta, "Action Plan",
        ACTION_PLAN_HEADERS, plan_rows,
        [16, 10, W_NUM, 20, W_TEXT, W_NUM, W_NUM, W_NUM, 20, 30, 34, 20, W_TEXT],
        wrap_columns=frozenset({5, 10, 11, 13}),
        color_column=7, color_map=SEVERITY_COLORS,
        empty_message="No actions were scheduled for this run.",
    )

    gantt_headers = ["Task #", "Action"] + [f"W{week}" for week in range(1, 17)]
    sheet = _add_sheet(
        workbook, meta, "Gantt",
        gantt_headers,
        [[action.get("id"), action.get("action")] + [""] * 16 for action in actions],
        [W_NUM, W_TEXT] + [W_WEEK] * 16,
        wrap_columns=frozenset({2}),
        empty_message="No actions were scheduled for this run.",
    )
    bar_fill = _fill(ACCENT)
    for row_offset, action in enumerate(actions):
        week = action.get("week")
        if not isinstance(week, int):
            continue
        start = max(1, min(16, week))
        end_raw = action.get("week_end")
        end = max(start, min(16, end_raw)) if isinstance(end_raw, int) else start
        for week_number in range(start, end + 1):
            sheet.cell(row=5 + row_offset, column=2 + week_number).fill = bar_fill

    _add_methodology(
        workbook, meta, data,
        "The canonical 16-week action sequence compiled from evidence-linked findings. "
        "Categories join each task to its source finding; unmatched tasks are General.",
    )
    return _save(workbook, path)


# --------------------------------------------------------------------------- on-page


def _onpage_workbook(data: dict[str, Any], path: Path, *, kind: str) -> Path:
    proposals = _proposals(data)
    workbook = _new_workbook()

    if kind == "title":
        meta = _meta(data, "Title Tag Optimizations")
        sheet_name = "Title Tags"
        current_key, proposed_key, rationale_key = (
            "current_title", "proposed_title", "title_rationale"
        )
        label = "Title"
        measured = (
            "Current versus proposed title tags for every reviewed page. A proposal that "
            "matches the current value is reported as 'No change required' rather than "
            "shipped as work."
        )
    elif kind == "meta":
        meta = _meta(data, "Meta Description Optimizations")
        sheet_name = "Meta Descriptions"
        current_key, proposed_key, rationale_key = (
            "current_meta", "proposed_meta", "meta_rationale"
        )
        label = "Meta Description"
        measured = (
            "Current versus proposed meta descriptions for every reviewed page. "
            "Unchanged proposals are reported as 'No change required'."
        )
    else:
        meta = _meta(data, "H1 Tags")
        sheet_name = "H1 Tags"
        current_key, proposed_key, rationale_key = (
            "current_h1", "proposed_h1", "h1_rationale"
        )
        label = "H1"
        measured = (
            "Current versus proposed H1 headings for every reviewed page. Unchanged "
            "proposals are reported as 'No change required' — an identical 'optimised' "
            "H1 is not a deliverable."
        )

    rows: list[list[Any]] = []
    changed = 0
    for proposal in proposals:
        current = proposal.get(current_key)
        rendered, status = _proposal_state(current, proposal.get(proposed_key))
        if status == "Proposed":
            changed += 1
        rows.append([
            proposal.get("page_id"),
            proposal.get("url"),
            proposal.get("page_type"),
            current,
            _text_length(current),
            rendered,
            _text_length(rendered) if status == "Proposed" else "—",
            status,
            proposal.get(rationale_key) or proposal.get("title_rationale") or UNAVAILABLE,
            proposal.get("target_keyword"),
            proposal.get("target_volume"),
            proposal.get("source"),
            proposal.get("approval_status"),
            proposal.get("evidence_ids"),
        ])
    _add_sheet(
        workbook, meta, sheet_name,
        ["Page ID", "URL", "Page Type", f"Current {label}", "Current Length",
         f"Proposed {label}", "Proposed Length", "Status", "Rationale",
         "Target Keyword", "Target Volume", "Source", "Approval Status",
         "Evidence IDs"],
        rows,
        [W_NUM, W_URL, W_LABEL, W_TEXT, W_NUM, W_TEXT, W_NUM, W_LABEL, W_TEXT,
         W_TEXT, W_NUM, W_LABEL, 26, 22],
        wrap_columns=frozenset({4, 6, 9}),
        color_column=8, color_map=ISSUE_COLORS,
        empty_message="No on-page proposals were produced for this run.",
        note=(
            f"{changed} of {len(rows)} reviewed pages carry a substantive change. "
            "Every proposal awaits editorial approval before deployment."
            if rows else None
        ),
    )
    _add_methodology(
        workbook, meta, data, measured,
        extra=[
            ["Rows reviewed", len(rows)],
            ["Substantive changes proposed", changed],
            ["Unchanged (no change required)", len(rows) - changed],
            [
                "Anti-pattern control",
                "A proposal byte-identical to the current value is never presented as an "
                "optimisation. This is the defect that made the previous package's H1 "
                "sheet unusable.",
            ],
        ],
    )
    return _save(workbook, path)


def _internal_link_map(data: dict[str, Any], path: Path) -> Path:
    meta = _meta(data, "Internal Link Map")
    deployment = data.get("deployment") or {}
    candidates = list(deployment.get("internal_link_candidates") or [])
    workbook = _new_workbook()

    _add_sheet(
        workbook, meta, "Internal Link Map",
        ["Source", "Target", "Anchor", "Rationale", "Link Type", "Observed",
         "Approval", "Evidence IDs"],
        [
            [candidate.get("source_url"), candidate.get("target_url"),
             candidate.get("anchor"), candidate.get("rationale"),
             candidate.get("link_type"), candidate.get("observed_status"),
             candidate.get("approval_status"), candidate.get("evidence_ids")]
            for candidate in candidates
        ],
        [W_URL, W_URL, W_TEXT, W_TEXT, 24, W_LABEL, 22, 22],
        wrap_columns=frozenset({3, 4}),
        empty_message="No internal link candidates cleared evidence checks during this run.",
    )
    hub_counts: Counter[str] = Counter(
        str(candidate.get("target_url")) for candidate in candidates
    )
    observed = {
        str(page.get("normalized_url")): page.get("internal_links")
        for page in _pages(data)
    }
    _add_sheet(
        workbook, meta, "Target Concentration",
        ["Target URL", "Proposed Inbound Links", "Currently Observed Internal Links"],
        [
            [url, count, observed.get(url)]
            for url, count in sorted(hub_counts.items(), key=lambda item: -item[1])
        ],
        [W_URL, W_NUM + 8, W_NUM + 12],
        empty_message="No internal link candidates cleared evidence checks during this run.",
    )
    _add_methodology(
        workbook, meta, data,
        "Internal link candidates derived from crawl-observed relationships. Anchors and "
        "targets were observed on the approved domain; additions await approval.",
    )
    return _save(workbook, path)


# --------------------------------------------------------------------------- fixes


def _redirect_map(data: dict[str, Any], path: Path) -> Path:
    meta = _meta(data, "Redirect Map")
    pages = _pages(data)
    deployment = data.get("deployment") or {}
    workbook = _new_workbook()

    rows: list[list[Any]] = []
    for candidate in deployment.get("redirect_candidates") or []:
        rows.append([
            candidate.get("source_url") or candidate.get("url"),
            candidate.get("target_url") or candidate.get("proposed_target"),
            candidate.get("status_code"), candidate.get("redirect_type") or "301",
            candidate.get("reason") or "Deployment candidate",
            candidate.get("approval_status") or "review_required",
            candidate.get("evidence_id") or candidate.get("evidence_ids"),
        ])
    seen = {str(row[0]) for row in rows}
    for page in pages:
        status = page.get("status_code")
        url = str(page.get("normalized_url") or "")
        if not isinstance(status, int) or status < 400 or url in seen:
            continue
        rows.append([
            url, UNAVAILABLE, status, "301",
            "Observed 4xx/5xx response; a destination must be chosen by the site owner "
            "before this redirect can be deployed.",
            "review_required", page.get("evidence_id"),
        ])
    _add_sheet(
        workbook, meta, "Redirect Map",
        ["Source URL", "Proposed Target", "Observed Status", "Redirect Type", "Reason",
         "Approval Status", "Evidence"],
        rows, [W_URL, W_URL, W_NUM, W_LABEL, W_TEXT, 24, W_NUM],
        wrap_columns=frozenset({5}),
        empty_message="No error responses or redirect candidates were observed.",
    )
    _add_sheet(
        workbook, meta, "Observed Chains",
        ["URL", "Chain", "Hops", "Status", "Evidence ID"],
        [
            [page.get("normalized_url"),
             " → ".join(str(hop) for hop in page.get("redirect_chain") or []),
             len(page.get("redirect_chain") or []), page.get("status_code"),
             page.get("evidence_id")]
            for page in pages
            if len(page.get("redirect_chain") or []) > 1
        ],
        [W_URL, W_URL + 20, W_NUM, W_NUM, W_NUM],
        wrap_columns=frozenset({2}),
        empty_message="No multi-hop redirect chains were observed during the crawl.",
    )
    _add_methodology(
        workbook, meta, data,
        "One row per observed error response and per deployment redirect candidate. A "
        "target is stated only when the evidence supports it; otherwise the destination "
        "is left Unavailable for the site owner to decide.",
    )
    return _save(workbook, path)


def _canonical_coverage(data: dict[str, Any]) -> list[list[Any]]:
    pages = _pages(data)
    if not pages:
        return []
    counter: Counter[str] = Counter(_canonical_state(page) for page in pages)
    return [
        [state, count, round(count / len(pages), 3)]
        for state, count in sorted(counter.items())
    ]


def _canonical_fixes(data: dict[str, Any], path: Path) -> Path:
    meta = _meta(data, "Canonical Fixes")
    deployment = data.get("deployment") or {}
    workbook = _new_workbook()

    rows: list[list[Any]] = []
    for candidate in deployment.get("canonical_candidates") or []:
        rows.append([
            candidate.get("url") or candidate.get("source_url"),
            candidate.get("current_canonical") or candidate.get("canonical_url"),
            candidate.get("proposed_canonical") or candidate.get("target_url"),
            candidate.get("reason") or candidate.get("issue"),
            "Deployment candidate",
            candidate.get("approval_status"),
            candidate.get("evidence_id") or candidate.get("evidence_ids"),
        ])
    seen = {str(row[0]) for row in rows}
    for page in _pages(data):
        state = _canonical_state(page)
        url = str(page.get("normalized_url") or "")
        if state == "Match" or url in seen:
            continue
        rows.append([
            url, page.get("canonical_url"), url,
            f"Canonical {state.casefold()} observed in crawl", "Crawl observation",
            "review_required", page.get("evidence_id"),
        ])
    _add_sheet(
        workbook, meta, "Canonical Fixes",
        ["URL", "Current Canonical", "Proposed Canonical", "Issue", "Source",
         "Approval Status", "Evidence"],
        rows, [W_URL, W_URL, W_URL, W_TEXT, W_LABEL + 4, 24, W_NUM],
        wrap_columns=frozenset({4}),
        empty_message="All observed canonicals were consistent during the crawl window.",
    )
    _add_sheet(
        workbook, meta, "Canonical Coverage",
        ["State", "Pages", "Share"], _canonical_coverage(data),
        [W_LABEL, W_NUM, W_NUM],
        empty_message="No pages were captured during the crawl window.",
    )
    _add_methodology(
        workbook, meta, data,
        "Canonical declarations compared against normalized URLs, plus any deployment "
        "candidates. Every change requires administrator approval before implementation.",
    )
    return _save(workbook, path)


# --------------------------------------------------------------------------- link build


def _referring_domains_workbook(data: dict[str, Any], path: Path) -> Path:
    meta = _meta(data, "Referring Domains")
    backlinks = _backlinks(data)
    workbook = _new_workbook()
    _add_sheet(
        workbook, meta, "Referring Domains",
        ["Referring Domain", "Authority Score", "Backlinks", "Country", "First Seen",
         "Last Seen", "Disposition"],
        [
            [
                row.get("domain"), row.get("authority_score"), row.get("backlinks"),
                row.get("country"), row.get("first_seen"), row.get("last_seen"),
                "Retain — no toxicity claim is made without manual review",
            ]
            for row in (backlinks.get("referring_domains") or [])
            if isinstance(row, dict)
        ],
        [W_TEXT, W_NUM, W_NUM, W_NUM, W_DATE, W_DATE, W_TEXT],
        wrap_columns=frozenset({7}),
        empty_message=_unavailable_message(backlinks, "no backlink provider connected"),
    )
    _add_methodology(
        workbook, meta, data,
        "The referring-domain inventory exactly as returned by the connected provider. "
        "No domain is labelled toxic and no disavow file is generated from this data.",
    )
    return _save(workbook, path)


def _link_gap_workbook(data: dict[str, Any], path: Path) -> Path:
    meta = _meta(data, "Link Gap Opportunities")
    workbook = _new_workbook()
    _add_sheet(
        workbook, meta, "Link Gap Opportunities",
        ["Competitor", "Referring Domains", "Common Keywords", "Gap Keywords",
         "Opportunity", "Evidence IDs"],
        _link_gap_rows(data, _backlinks_available(data)),
        [W_TEXT, W_NUM, W_NUM, W_NUM, W_TEXT, 22],
        wrap_columns=frozenset({5}),
        empty_message=_unavailable_message(
            _backlinks(data), "no backlink provider connected"
        ),
    )
    _add_methodology(
        workbook, meta, data,
        "Competitor link gaps as reported by the connected provider. Outreach targets "
        "must be manually qualified; no contact details are generated here.",
    )
    return _save(workbook, path)


# --------------------------------------------------------------------------- QC


def _qc_report(data: dict[str, Any], path: Path) -> Path:
    meta = _meta(data, "Quality Control Report")
    qa = data.get("qa") or {}
    workbook = _new_workbook()

    _add_sheet(
        workbook, meta, "Release Gates",
        ["Gate", "Status", "Critical Failures", "High Failures", "Evidence"],
        [
            [gate.get("name"), gate.get("status"), gate.get("critical_failures"),
             gate.get("high_failures"), gate.get("evidence")]
            for gate in qa.get("gates") or []
        ],
        [40, W_LABEL, W_NUM, W_NUM, W_TEXT],
        wrap_columns=frozenset({5}),
        empty_message="No release gates were recorded for this run.",
    )
    _add_sheet(
        workbook, meta, "Reconciliation",
        ["Measure", "Canonical", "Package", "Result"],
        [
            [item.get("measure"), item.get("canonical"), item.get("package"),
             item.get("result")]
            for item in qa.get("reconciliation") or []
        ],
        [46, W_NUM + 8, W_NUM + 8, 30],
        empty_message="No reconciliation rows were recorded for this run.",
    )
    _add_sheet(
        workbook, meta, "Source Coverage",
        ["Source ID", "Source", "Kind", "Status", "Coverage", "Scope",
         "Unavailable Reason"],
        [
            [source.get("id"), source.get("label"), source.get("kind"),
             source.get("status"), source.get("coverage"), source.get("scope"),
             source.get("unavailable_reason") or "Not applicable"]
            for source in data.get("sources") or []
        ],
        [W_LABEL, W_TEXT, W_LABEL, W_LABEL, W_NUM, W_TEXT, W_TEXT],
        wrap_columns=frozenset({6, 7}),
        empty_message="No sources were registered for this run.",
    )

    integrity = data.get("crawl_integrity") or {}
    _add_sheet(
        workbook, meta, "Crawl Integrity",
        ["Measure", "Value"],
        [
            ["Crawl status", integrity.get("status") or UNAVAILABLE],
            ["Pages fetched", integrity.get("fetched_pages")],
            ["Challenged pages", integrity.get("challenged_pages")],
            ["Challenge share", integrity.get("challenge_share")],
            ["Rate limited pages", integrity.get("rate_limited_pages")],
            ["Quarantined URLs", integrity.get("quarantined_urls")],
            ["Note", integrity.get("note")],
        ] if integrity else [],
        [34, 70], wrap_columns=frozenset({2}),
        empty_message=f"{UNAVAILABLE} — crawl integrity telemetry was not recorded.",
    )
    _add_sheet(
        workbook, meta, "Limitations",
        ["#", "Limitation"],
        [[index, text] for index, text in enumerate(data.get("limitations") or [], 1)],
        [6, 100], wrap_columns=frozenset({2}),
        empty_message="No limitations were recorded for this run.",
    )
    _add_methodology(
        workbook, meta, data,
        "The release record for this package: gate outcomes, count reconciliation between "
        "the canonical data and the rendered files, source coverage, crawl integrity "
        "telemetry and the declared limitations.",
    )
    return _save(workbook, path)


# --------------------------------------------------------------------------- entry


def render_workbooks(data: dict, package_root: Path) -> list[Path]:
    """Render every XLSX workbook of the package tree. Returns written paths.

    Member paths come from :mod:`exporters.paths`, the single source of truth for
    the package tree — this module holds no path literal of its own.
    """
    root = Path(package_root)

    written: list[Path] = [
        _technical_audit(data, root / tree.TECHNICAL_AUDIT_XLSX),
        _content_audit(data, root / tree.CONTENT_AUDIT_XLSX),
        _backlink_audit(data, root / tree.BACKLINK_AUDIT_XLSX),
        _competitor_landscape(data, root / tree.COMPETITOR_LANDSCAPE_XLSX),
        _geo_aeo(data, root / tree.GEO_AEO_XLSX),
        _cro_ux(data, root / tree.CRO_UX_XLSX),
        _tracking_audit(data, root / tree.TRACKING_AUDIT_XLSX),
        _baseline_performance(data, root / tree.BASELINE_PERFORMANCE_XLSX),
        _master_keyword_universe(data, root / tree.MASTER_KEYWORD_UNIVERSE_XLSX),
        _content_gap_analysis(data, root / tree.CONTENT_GAP_XLSX),
        _content_strategy(data, root / tree.CONTENT_STRATEGY_XLSX),
        _url_architecture(data, root / tree.URL_ARCHITECTURE_XLSX),
        _cannibalization_plan(data, root / tree.CANNIBALIZATION_XLSX),
        _action_plan(data, root / tree.ACTION_PLAN_XLSX),
        _onpage_workbook(data, root / tree.TITLE_TAG_XLSX, kind="title"),
        _onpage_workbook(data, root / tree.META_DESCRIPTION_XLSX, kind="meta"),
        _onpage_workbook(data, root / tree.H1_TAGS_XLSX, kind="h1"),
        _internal_link_map(data, root / tree.INTERNAL_LINK_MAP_XLSX),
        _redirect_map(data, root / tree.REDIRECT_MAP_XLSX),
        _canonical_fixes(data, root / tree.CANONICAL_FIXES_XLSX),
        _qc_report(data, root / tree.QC_REPORT_XLSX),
    ]
    if _is_ecommerce(data):
        written.append(_ecommerce_audit(data, root / tree.ECOMMERCE_AUDIT_XLSX))
    if _is_local(data):
        written.append(_gbp_local(data, root / tree.GBP_LOCAL_XLSX))
    if _backlinks_available(data):
        written.append(_referring_domains_workbook(data, root / tree.REFERRING_DOMAINS_XLSX))
        written.append(_link_gap_workbook(data, root / tree.LINK_GAP_XLSX))
    return written
