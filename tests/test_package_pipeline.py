"""End-to-end tests for the audit deliverable package pipeline (Track D)."""

from __future__ import annotations

import hashlib
import json
import zipfile
from io import BytesIO
from pathlib import Path

import pytest

from app import views
from app.domain.constants import RunState, StageStatus, UserRole
from app.domain.models import (
    Artifact,
    AuditRun,
    Client,
    PackageManifest,
    Project,
    RunStage,
    User,
)
from app.domain.storage import open_verified_artifact
from audit_engine.crawler import CrawledPage, CrawlResult
from audit_engine.tasks import run_website_audit
from exporters import paths as tree
from exporters.tasks import build_audit_package

REPO_ROOT = Path(__file__).resolve().parents[1]
PROGRESS_KEYS = {
    "state", "label", "percent", "pages", "findings",
    "recommendations", "active", "failed", "message",
}


class MemoryStorage:
    def __init__(self) -> None:
        self.objects: dict[str, bytes] = {}

    def exists(self, name: str) -> bool:
        return name in self.objects

    def save(self, name: str, content) -> str:
        self.objects[name] = content.read()
        return name

    def open(self, name: str, mode: str = "rb") -> BytesIO:
        if mode != "rb":
            raise ValueError("Test storage is read-only after save")
        return BytesIO(self.objects[name])


@pytest.fixture
def storage(monkeypatch) -> MemoryStorage:
    memory = MemoryStorage()
    monkeypatch.setattr("app.domain.storage.default_storage", memory)
    return memory


def _seed_run(*, client_name: str = "Aurora Test Homewares", slug: str = "package-client") -> AuditRun:
    user = User.objects.create_user(
        username=f"{slug}-admin",
        password="Package-test-password-9911!",  # noqa: S106 - test credential
        role=UserRole.AGENCY_ADMIN,
        must_change_password=False,
    )
    client = Client.objects.create(name=client_name, slug=slug)
    project = Project.objects.create(
        client=client,
        name=f"{client_name} SEO",
        slug=slug,
        primary_domain="example.com.au",
        approved_domains=["example.com.au"],
        business_type=Project.BusinessType.ECOMMERCE,
    )
    return AuditRun.objects.create(
        project=project,
        profile="quick",
        idempotency_key=f"{slug}-run",
        rule_version="1.0.0",
        created_by=user,
    )


def _run_real_audit(run: AuditRun, monkeypatch) -> None:
    result = CrawlResult(
        pages=(
            CrawledPage(
                requested_url="https://example.com.au/",
                final_url="https://example.com.au/",
                status_code=200, content_type="text/html", body_sha256="a" * 64,
                title="Example home", meta_description=None, h1=("Welcome",),
                canonical_url=None, robots_directives=(),
                links=("https://example.com.au/shop",),
                redirect_chain=("https://example.com.au/",),
            ),
            CrawledPage(
                requested_url="https://example.com.au/shop",
                final_url="https://example.com.au/shop",
                status_code=200, content_type="text/html", body_sha256="b" * 64,
                title=None, meta_description=None, h1=(),
                canonical_url=None, robots_directives=(), links=(),
                redirect_chain=("https://example.com.au/shop",),
            ),
        ),
        failures=(), discovered_count=2, stopped_reason="queue_exhausted",
    )

    class FakeCrawler:
        def __init__(self, config):
            self.config = config

        def crawl(self, seeds):
            return result

    monkeypatch.setattr("audit_engine.tasks.BoundedCrawler", FakeCrawler)
    output = run_website_audit.run(str(run.pk))
    assert output["state"] == RunState.GATE_1_REVIEW


@pytest.mark.django_db
@pytest.mark.render
def test_package_pipeline_builds_verifiable_zip_manifest_and_progress(
    monkeypatch, settings, tmp_path, storage
):
    pytest.importorskip("exporters.run_data", reason="Track A exporters.run_data has not landed yet")
    pytest.importorskip(
        "exporters.xlsx_workbooks", reason="Track B exporters.xlsx_workbooks has not landed yet"
    )
    pytest.importorskip("exporters.pptx_deck", reason="Track C exporters.pptx_deck has not landed yet")
    pytest.importorskip(
        "exporters.markdown_summary", reason="Track C exporters.markdown_summary has not landed yet"
    )
    settings.AUTO_BUILD_PACKAGE = False

    run = _seed_run()
    _run_real_audit(run, monkeypatch)
    run.refresh_from_db()

    result = build_audit_package.run(str(run.pk))
    assert result.get("failed") is not True, f"Package build failed: {result.get('error')}"
    assert result["idempotent"] is False
    assert result["artifact_id"]

    artifact = Artifact.objects.get(pk=result["artifact_id"])
    assert artifact.artifact_type == "package"
    assert artifact.format == "zip"
    assert artifact.sha256 == result["sha256"]
    assert artifact.metadata["run_version"] == run.version

    with open_verified_artifact(artifact) as stream:
        payload = stream.read()
    archive = zipfile.ZipFile(BytesIO(payload))
    names = archive.namelist()
    package_name = artifact.metadata["package_name"]
    assert package_name.startswith("Aurora_Test_Homewares_SEO_Audit_Package_")
    assert all(name.startswith(package_name + "/") for name in names)
    assert f"{package_name}/{tree.SUMMARY_MARKDOWN}" in names
    assert f"{package_name}/{tree.PACKAGE_MANIFEST_JSON}" in names
    assert f"{package_name}/{tree.CHECKSUMS_SHA256}" in names

    top_folders = {name.split("/")[1] for name in names if name.count("/") >= 2}
    for expected in (
        tree.AUDIT_REPORTS,
        tree.STRATEGY_DOCUMENTS,
        tree.ACTION_PLAN,
        tree.IMPLEMENTATION,
        tree.QA,
        tree.SLIDE_DECK,
    ):
        assert expected in top_folders, f"Missing package folder: {expected}"
    manifest_doc = json.loads(archive.read(f"{package_name}/{tree.PACKAGE_MANIFEST_JSON}"))
    if any(entry["artifact_type"] == "content_brief" for entry in manifest_doc["files"]):
        assert tree.SEO_CONTENT in top_folders

    # The V18 taxonomy is the contract: legacy folder names must be gone.
    assert "06_QA_and_Manifest" not in top_folders
    assert "05_Content" not in top_folders
    assert "07_Executive_Deck" not in top_folders

    members = {name.split("/", 1)[1] for name in names if "/" in name}

    # No byte-identical payloads anywhere in the package (V18 shipped 24; we ship none).
    digests: dict[str, str] = {}
    for name in names:
        digest = hashlib.sha256(archive.read(name)).hexdigest()
        assert digest not in digests, f"Duplicate payload: {name} == {digests[digest]}"
        digests[digest] = name

    # Every PPTX ships a PDF sibling, declared in the manifest as a derivative.
    manifest_by_path = {entry["path"]: entry for entry in manifest_doc["files"]}
    pptx_members = [name for name in members if name.endswith(".pptx")]
    assert pptx_members
    for member in pptx_members:
        sibling = member[: -len(".pptx")] + ".pdf"
        assert sibling in members, f"PPTX without a sibling PDF: {member}"
        assert manifest_by_path[sibling]["derivative_of"] == member

    # Deployment CSVs carry at least their header row and are declared in the manifest.
    assert f"{package_name}/{tree.REDIRECT_MAP_CSV}" in names
    assert tree.REDIRECT_MAP_CSV in manifest_by_path

    from openpyxl import load_workbook
    from pptx import Presentation
    from pypdf import PdfReader

    xlsx_member = next(name for name in names if name.endswith(".xlsx"))
    pdf_member = f"{package_name}/{tree.ENTERPRISE_AUDIT_PDF}"
    pptx_member = f"{package_name}/{tree.DECK_PPTX}"
    assert pdf_member in names
    assert pptx_member in names
    assert f"{package_name}/{tree.DECK_HTML}" in names
    for member in (xlsx_member, pdf_member, pptx_member):
        (tmp_path / Path(member).name).write_bytes(archive.read(member))
    assert load_workbook(tmp_path / Path(xlsx_member).name).sheetnames
    assert len(Presentation(tmp_path / Path(pptx_member).name).slides) >= 1
    assert len(PdfReader(tmp_path / Path(pdf_member).name).pages) >= 1

    manifest_row = PackageManifest.objects.get(run=run, version=run.version)
    assert manifest_row.package_artifact_id == artifact.pk
    assert manifest_row.package_sha256 == artifact.sha256

    progress = views._audit_progress_payload(run)
    assert set(progress) >= PROGRESS_KEYS
    assert progress["percent"] == 100
    assert progress["label"] == "Audit complete — package ready to download"
    assert progress["active"] is False

    repeat = build_audit_package.run(str(run.pk))
    assert repeat["idempotent"] is True
    assert repeat["artifact_id"] == str(artifact.pk)
    assert Artifact.objects.filter(run=run, artifact_type="package").count() == 1


@pytest.mark.django_db
def test_failed_package_build_returns_failure_dict_and_flags_progress(monkeypatch, storage):
    run = _seed_run(slug="package-fail")
    run.state = RunState.GATE_1_REVIEW
    run.save(update_fields=["state", "updated_at"])

    def explode(run_arg, *, actor=None, progress=None):
        if progress is not None:
            progress("Rendering audit workbooks")
        raise RuntimeError("boom: renderer unavailable")

    monkeypatch.setattr("exporters.tasks.build_package_for_run", explode)
    result = build_audit_package.run(str(run.pk))

    assert result["idempotent"] is False
    assert result.get("failed") is True
    assert result["artifact_id"] is None
    assert "boom" in result["error"]

    stage = RunStage.objects.get(run=run, name="packaging")
    assert stage.status == StageStatus.FAILED
    assert "boom" in stage.error_summary
    assert Artifact.objects.filter(run=run, artifact_type="package").count() == 0

    progress = views._audit_progress_payload(run)
    assert set(progress) >= PROGRESS_KEYS
    assert progress["label"] == "Audit ready — package build needs attention"
    assert progress["active"] is False
    assert progress["failed"] is False
    assert "boom" in progress["message"]


@pytest.mark.django_db
def test_progress_payload_reports_packaging_stage_while_running():
    run = _seed_run(slug="package-progress")
    run.state = RunState.GATE_1_REVIEW
    run.save(update_fields=["state", "updated_at"])
    RunStage.objects.create(
        run=run,
        name="packaging",
        sequence=30,
        status=StageStatus.RUNNING,
        checkpoint={"message": "Rendering audit workbooks"},
    )

    progress = views._audit_progress_payload(run)

    assert set(progress) >= PROGRESS_KEYS
    assert progress["percent"] == 90
    assert progress["label"] == "Building the deliverable package"
    assert progress["active"] is True
    assert progress["message"] == "Rendering audit workbooks"


@pytest.mark.django_db
def test_manual_package_build_endpoint_queues_and_throttles(monkeypatch, client):
    import exporters.tasks as exporter_tasks

    run = _seed_run(slug="package-manual")
    run.state = RunState.GATE_1_REVIEW
    run.save(update_fields=["state", "updated_at"])

    queued: list[tuple[str, str | None]] = []
    monkeypatch.setattr(
        exporter_tasks.build_audit_package,
        "delay",
        lambda run_id, actor_id=None: queued.append((run_id, actor_id)),
    )

    assert client.login(username="package-manual-admin", password="Package-test-password-9911!")  # noqa: S106 - test credential
    url = f"/projects/{run.project_id}/package/build/"

    response = client.post(url, secure=True)
    assert response.status_code == 302
    assert queued and queued[0][0] == str(run.pk)
    stage = RunStage.objects.get(run=run, name="packaging")
    assert (stage.checkpoint or {}).get("dispatched_at")

    # Auto-redispatch (non-forced) is throttled by the fresh dispatched_at stamp.
    run.refresh_from_db()
    assert views._dispatch_package_build(run) is False
    assert len(queued) == 1

    # The forced manual endpoint may re-queue.
    response = client.post(url, secure=True)
    assert response.status_code == 302
    assert len(queued) == 2

    # A run that is still auditing cannot queue a package.
    run.state = RunState.AUDITING
    run.save(update_fields=["state", "updated_at"])
    response = client.post(url, secure=True)
    assert response.status_code == 302
    assert len(queued) == 2


@pytest.mark.django_db
def test_progress_poll_autodispatches_missing_package(monkeypatch, client, settings):
    import exporters.tasks as exporter_tasks

    settings.AUTO_BUILD_PACKAGE = True
    run = _seed_run(slug="package-auto")
    run.state = RunState.GATE_1_REVIEW
    run.save(update_fields=["state", "updated_at"])

    queued: list[str] = []
    monkeypatch.setattr(
        exporter_tasks.build_audit_package,
        "delay",
        lambda run_id, actor_id=None: queued.append(run_id),
    )

    assert client.login(username="package-auto-admin", password="Package-test-password-9911!")  # noqa: S106 - test credential
    response = client.get(f"/projects/{run.project_id}/progress/", secure=True)
    assert response.status_code == 200
    assert queued == [str(run.pk)]

    # Second poll inside the throttle window must not double-queue.
    response = client.get(f"/projects/{run.project_id}/progress/", secure=True)
    assert response.status_code == 200
    assert queued == [str(run.pk)]


def _write(root: Path, relative: str, content: bytes) -> Path:
    output = root.joinpath(*relative.split("/"))
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_bytes(content)
    return output


def test_verify_tree_flags_a_pptx_without_a_pdf_sibling(tmp_path):
    from exporters.run_package import _verify_tree

    root = tmp_path / "pkg"
    _write(root, tree.SUMMARY_MARKDOWN, b"# summary\n")
    _write(root, tree.DECK_PPTX, b"not-a-real-pptx")

    failures, _counts, _total = _verify_tree(root, {"client": {"name": "Aurora"}})

    assert any("without a sibling PDF" in failure for failure in failures)


def test_verify_tree_flags_a_header_only_redirect_map_when_404s_exist(tmp_path):
    from exporters.run_package import _verify_tree

    root = tmp_path / "pkg"
    _write(root, tree.SUMMARY_MARKDOWN, b"# summary\n")
    _write(root, tree.REDIRECT_MAP_CSV, b"source_url,target_url,status_code\n")
    data = {
        "client": {"name": "Aurora"},
        "deployment": {"redirect_candidates": []},
        "pages": [{"status_code": 404}, {"status_code": 410}, {"status_code": 200}],
    }

    failures, _counts, _total = _verify_tree(root, data)

    assert any("Header-only deployment CSV" in failure for failure in failures)

    # With rows present the same tree verifies cleanly.
    _write(
        root,
        tree.REDIRECT_MAP_CSV,
        b"source_url,target_url,status_code\nhttps://a/,https://b/,404\n",
    )
    failures, _counts, _total = _verify_tree(root, data)
    assert not [failure for failure in failures if "Header-only" in failure]


def _tree_data() -> dict:
    return {
        "client": {"name": "Aurora Test Homewares", "domain": "example.com.au"},
        "project": {"business_profile": "ecommerce"},
        "run": {
            "id": "RUN-AURORA-TREE",
            "version": 1,
            "rule_version": "1.0.0",
            "evidence_as_of": "2026-07-20",
            "captured_at": "2026-07-20T00:00:00+00:00",
        },
        "categories": [
            {"key": "technical", "category": "Technical", "score": 61, "coverage": 0.8,
             "weight": 0.3, "score_reason": None, "unavailable_reason": None},
        ],
        "findings": [{"id": "F-1", "category": "technical", "severity": "High"}],
        "actions": [{"id": "ACT-1", "phase": "Foundation", "week": 1, "action": "Fix redirects"}],
        "sources": [{"id": "SRC-1", "label": "Crawl", "status": "available"}],
        "evidence": [{"id": "EV-1", "source_id": "SRC-1", "evidence_type": "crawl"}],
        "generation_ledger": [{"id": "GEN-1", "task": "summary", "status": "skipped"}],
        "deployment": {
            "redirect_candidates": [
                {"source_url": "https://example.com.au/old", "target_url": "https://example.com.au/",
                 "status_code": 404, "included_in_deployment": True},
            ],
            "robots": {"recommendation": "Keep the current directives."},
        },
        "pages": [{"status_code": 404}],
        "backlinks": {"status": "unavailable", "unavailable_reason": "No provider connected."},
        "crawl_integrity": {"status": "clean"},
        "content_assets": [],
        "qa": {"reconciliation": []},
        "limitations": [],
    }


def test_package_writers_produce_the_v18_tree_with_a_clean_manifest(tmp_path):
    from exporters.manifest import PackageManifest, resolve_derivative_of
    from exporters.run_package import (
        _prune_unavailable_folders,
        _verify_tree,
        _write_csvs,
        _write_markdown,
        _write_robots_notes,
        _write_schema_files,
    )

    data = _tree_data()
    root = tmp_path / "Aurora_Test_Homewares_SEO_Audit_Package_20260720"
    root.mkdir()
    tree.ensure_folders(root)
    _write_markdown(data, root)
    _write_csvs(data, root)
    _write_robots_notes(data, root)
    _write_schema_files(data, root, {"business_summary": "Homewares retailer."})
    _prune_unavailable_folders(data, root)

    written = {path.relative_to(root).as_posix() for path in root.rglob("*") if path.is_file()}
    for expected in (
        tree.SUMMARY_MARKDOWN,
        tree.ACTION_PLAN_CSV,
        tree.REDIRECT_MAP_CSV,
        tree.ROBOTS_RECOMMENDATIONS_TXT,
        tree.SCHEMA_ORGANIZATION_JSON,
        tree.SCHEMA_PRODUCT_TEMPLATE_JSON,
        tree.SOURCE_COVERAGE_CSV,
        tree.CHANGE_LOG_CSV,
        tree.EVIDENCE_INDEX_CSV,
        tree.ISSUE_REGISTER_CSV,
    ):
        assert expected in written, f"Missing {expected}"
    # Backlinks were unavailable, so the link-building folder is absent, not empty.
    assert not (root / "04_Implementation_Deliverables" / "Link_Building").exists()
    assert not any(path.as_posix().startswith("06_QA_and_Manifest") for path in map(Path, written))

    failures, counts, total = _verify_tree(root, data)
    assert failures == []
    assert total == len(written)
    assert counts[".csv"] >= 6

    manifest = PackageManifest(
        package_id=root.name, project_id="P", run_id="R", rule_version="1.0.0",
        evidence_as_of="2026-07-20",
    )
    available = set(written)
    for relative in sorted(written):
        artifact_type, approval_state = tree.entry_profile(relative)
        manifest.add_file(
            root,
            root.joinpath(*relative.split("/")),
            artifact_type=artifact_type,
            title=relative,
            approval_state=approval_state,
            derivative_of=resolve_derivative_of(relative, available),
        )
    manifest.assert_integrity(root)  # no duplicates, full coverage, valid derivatives


def test_markdown_summary_renders_the_new_market_and_keyword_sections():
    from exporters.markdown_summary import render_markdown

    data = _tree_data()
    data["executive_summary"] = "Crawl evidence supports a staged technical clean-up."
    data["market"] = {
        "status": "available", "provider": "semrush", "database": "au",
        "fetched_at": "2026-07-20T00:00:00+00:00",
        "domain": {"organic_keywords": 412, "organic_traffic": 3100, "organic_cost": 5200.0,
                   "adwords_keywords": 0, "authority_score": 24, "backlinks_total": 890,
                   "referring_domains": 61},
    }
    data["performance_vs_competitors"] = {
        "status": "available",
        "metrics": [{"metric": "Organic keywords", "client": 412, "competitor_median": 980,
                     "best_competitor": "rival.com.au", "best_value": 2100,
                     "position": "behind", "note": ""}],
        "summary": "The site trails the competitor median on organic keyword coverage.",
    }
    data["keywords"] = [
        {"id": "KW-0001", "phrase": "linen bedding", "position": 14, "search_volume": 2400,
         "cpc": 1.85, "funnel_stage": "BOFU", "landing_url": "https://example.com.au/bedding"},
        {"id": "KW-0002", "phrase": "cotton throw", "position": None, "search_volume": None,
         "cpc": None, "funnel_stage": None, "landing_url": None},
    ]
    data["crawl_integrity"] = {
        "status": "degraded", "fetched_pages": 180, "challenged_pages": 20,
        "challenge_share": 0.1, "rate_limited_pages": 4,
        "quarantined_urls": ["https://example.com.au/blocked"],
        "note": "Challenged pages were excluded from scoring.",
    }

    markdown = render_markdown(data)

    assert "## Market position" in markdown
    assert "## How you compare" in markdown
    assert "## Keyword opportunities" in markdown
    assert "## Crawl integrity" in markdown
    assert "linen bedding" in markdown
    assert "rival.com.au" in markdown
    assert "Executive_Deck.html" in markdown
    assert "Master_Keyword_Universe.xlsx" in markdown
    assert "Link_Building/ omitted" in markdown
    assert "None" not in markdown
    assert "06_QA_and_Manifest" not in markdown


def test_markdown_summary_is_honest_when_market_data_is_unavailable():
    from exporters.markdown_summary import render_markdown

    data = _tree_data()
    data["market"] = {"status": "unavailable", "unavailable_reason": "No provider key configured."}
    data["performance_vs_competitors"] = {
        "status": "unavailable", "unavailable_reason": "No competitor set was retrieved.",
    }
    data["keywords"] = []

    markdown = render_markdown(data)

    assert "No provider key configured." in markdown
    assert "No competitor set was retrieved." in markdown
    assert "no keyword provider returned measured search volumes" in markdown
    # An unavailable crawl-integrity status stays silent rather than implying a problem.
    assert "## Crawl integrity" not in markdown
    assert "None" not in markdown


def test_entry_profile_uses_the_most_specific_folder_rule():
    assert tree.entry_profile(tree.SUMMARY_MARKDOWN) == ("summary_markdown", "approved")
    assert tree.entry_profile(tree.TITLE_TAG_XLSX) == (
        "onpage_proposal",
        "withheld_pending_editorial_review",
    )
    assert tree.entry_profile(tree.REDIRECT_MAP_CSV) == (
        "deployment_asset",
        "withheld_pending_approval",
    )
    assert tree.entry_profile(tree.SCHEMA_ORGANIZATION_JSON)[0] == "schema_template"
    assert tree.entry_profile(f"{tree.SEO_CONTENT}/CONTENT-01_guide.docx") == (
        "content_brief",
        "withheld_pending_human_approval",
    )


def test_derivative_resolution_prefers_the_richest_sibling_source():
    from exporters.manifest import resolve_derivative_of

    available = {
        tree.DECK_PPTX, tree.DECK_PDF, tree.DECK_HTML,
        tree.SEO_STRATEGY_DOCX, tree.SEO_STRATEGY_PDF,
        tree.ACTION_PLAN_XLSX, tree.ACTION_PLAN_CSV, tree.ACTION_PLAN_PDF,
        tree.ENTERPRISE_AUDIT_PDF,
    }
    assert resolve_derivative_of(tree.DECK_PDF, available) == tree.DECK_PPTX
    assert resolve_derivative_of(tree.DECK_HTML, available) == tree.DECK_PPTX
    assert resolve_derivative_of(tree.SEO_STRATEGY_PDF, available) == tree.SEO_STRATEGY_DOCX
    assert resolve_derivative_of(tree.ACTION_PLAN_CSV, available) == tree.ACTION_PLAN_XLSX
    # A standalone PDF never claims a source it was not rendered from.
    assert resolve_derivative_of(tree.ENTERPRISE_AUDIT_PDF, available) is None
    assert resolve_derivative_of(tree.DECK_PPTX, available) is None


@pytest.mark.render
def test_docx_strategy_title_and_subtitle_follow_the_client_name(tmp_path):
    from docx import Document

    from exporters.docx_reports import DOCXReportBuilder

    data = {
        "client": {"name": "Aurora Homewares", "domain": "aurorahomewares.com.au"},
        "run": {
            "id": "RUN-AURORA-TEST",
            "evidence_as_of": "2026-07-17",
            "overall_score_reason": "Withheld because weighted evidence coverage is below 70%.",
        },
        "executive_summary": "Observed crawl evidence supports a staged technical clean-up.",
        "strategy_sections": [],
        "opportunities": [],
        "measurement_plan": [],
        "sources": [],
        "limitations": ["Private analytics remain unavailable until a connection is approved."],
    }
    output = DOCXReportBuilder(tmp_path).strategy_report(data, tmp_path / "strategy.docx")
    document = Document(output)

    assert document.core_properties.title == "Aurora Homewares Enterprise SEO Strategy"
    text = "\n".join(paragraph.text for paragraph in document.paragraphs)
    assert "Aurora Homewares" in text
    assert "Kakawa" not in text


@pytest.mark.deployment
def test_embedded_worker_consumes_the_render_queue():
    entrypoint = (REPO_ROOT / "deployment" / "entrypoint.sh").read_text(encoding="utf-8")
    web_block = entrypoint.split("web)", 1)[1].split(";;", 1)[0]
    assert "--queues=analysis,render" in web_block


@pytest.mark.deployment
def test_package_render_dependencies_are_declared():
    pyproject = (REPO_ROOT / "pyproject.toml").read_text(encoding="utf-8")
    dependency_block = pyproject.split("dependencies = [", 1)[1].split("\n]", 1)[0]
    assert '"openpyxl>=' in dependency_block
    assert '"python-pptx>=' in dependency_block


# ---------------------------------------------------------------------------
# Queue starvation guards: live audits must never wait behind backfills
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_backfill_lane_dispatches_to_render_queue(monkeypatch):
    import exporters.tasks as exporter_tasks

    run = _seed_run(slug="backfill-lane")
    run.state = RunState.GATE_1_REVIEW
    run.save(update_fields=["state", "updated_at"])

    lanes: list[tuple[str, dict]] = []
    monkeypatch.setattr(
        exporter_tasks.build_audit_package,
        "apply_async",
        lambda args=None, **kwargs: lanes.append(("apply_async", kwargs)),
    )
    monkeypatch.setattr(
        exporter_tasks.build_audit_package,
        "delay",
        lambda *args, **kwargs: lanes.append(("delay", kwargs)),
    )

    assert views._dispatch_package_build(run, lane="backfill") is True
    assert lanes == [("apply_async", {"queue": "render"})]

    # The normal pipeline lane keeps using the analysis queue via delay().
    stage = RunStage.objects.get(run=run, name="packaging")
    stage.checkpoint = {}
    stage.save(update_fields=["checkpoint", "updated_at"])
    assert views._dispatch_package_build(run) is True
    assert lanes[-1][0] == "delay"


@pytest.mark.django_db
def test_dashboard_backfill_waits_while_an_audit_is_live(monkeypatch, client, settings):
    import exporters.tasks as exporter_tasks

    settings.AUTO_BUILD_PACKAGE = True
    run = _seed_run(slug="backfill-guard")
    run.state = RunState.GATE_1_REVIEW
    run.save(update_fields=["state", "updated_at"])

    live_project = Project.objects.create(
        client=run.project.client,
        name="Live Crawl",
        slug="backfill-guard-live",
        primary_domain="live.example.com.au",
        approved_domains=["live.example.com.au"],
        business_type=Project.BusinessType.SERVICE,
    )
    live_run = AuditRun.objects.create(
        project=live_project,
        profile="quick",
        idempotency_key="backfill-guard-live-run",
        rule_version="1.0.0",
        state=RunState.COLLECTING,
    )

    queued: list[dict] = []
    monkeypatch.setattr(
        exporter_tasks.build_audit_package,
        "apply_async",
        lambda args=None, **kwargs: queued.append(kwargs),
    )
    monkeypatch.setattr(
        exporter_tasks.build_audit_package,
        "delay",
        lambda *args, **kwargs: queued.append({"lane": "delay"}),
    )

    assert client.login(username="backfill-guard-admin", password="Package-test-password-9911!")  # noqa: S106 - test credential

    # While another audit is genuinely running, no backfill may be queued.
    assert client.get("/", secure=True).status_code == 200
    assert queued == []

    # Once the live audit finishes, the same page load queues the rebuild
    # on the low-priority render lane.
    live_run.state = RunState.FAILED
    live_run.save(update_fields=["state", "updated_at"])
    assert client.get("/", secure=True).status_code == 200
    assert queued == [{"queue": "render"}]


@pytest.mark.django_db
def test_stalled_queued_run_is_rekicked_once(monkeypatch, client):
    from datetime import timedelta

    from django.core.cache import cache
    from django.utils import timezone

    import audit_engine.tasks as audit_tasks

    cache.clear()
    run = _seed_run(slug="queued-kick")
    AuditRun.objects.filter(pk=run.pk).update(
        updated_at=timezone.now() - timedelta(seconds=600)
    )

    kicked: list[str] = []
    monkeypatch.setattr(
        audit_tasks.run_website_audit, "delay", lambda run_id: kicked.append(run_id)
    )

    assert client.login(username="queued-kick-admin", password="Package-test-password-9911!")  # noqa: S106 - test credential
    url = f"/projects/{run.project_id}/progress/"

    assert client.get(url, secure=True).status_code == 200
    assert kicked == [str(run.pk)]

    # The cache lock stops the poll loop from flooding the queue.
    AuditRun.objects.filter(pk=run.pk).update(
        updated_at=timezone.now() - timedelta(seconds=600)
    )
    assert client.get(url, secure=True).status_code == 200
    assert kicked == [str(run.pk)]

    cache.clear()


@pytest.mark.django_db
def test_fresh_queued_run_is_not_rekicked(monkeypatch, client):
    from django.core.cache import cache

    import audit_engine.tasks as audit_tasks

    cache.clear()
    run = _seed_run(slug="queued-fresh")

    kicked: list[str] = []
    monkeypatch.setattr(
        audit_tasks.run_website_audit, "delay", lambda run_id: kicked.append(run_id)
    )

    assert client.login(username="queued-fresh-admin", password="Package-test-password-9911!")  # noqa: S106 - test credential
    response = client.get(f"/projects/{run.project_id}/progress/", secure=True)

    assert response.status_code == 200
    assert kicked == []


def test_worker_drains_analysis_before_render():
    from django.conf import settings as django_settings

    options = django_settings.CELERY_BROKER_TRANSPORT_OPTIONS
    assert options.get("queue_order_strategy") == "priority"

    entrypoint = (REPO_ROOT / "deployment" / "entrypoint.sh").read_text(encoding="utf-8")
    assert "--queues=analysis,render" in entrypoint


# ---------------------------------------------------------------------------
# Audit-task hardening: cancel wins, FAILED retry cleans up, resume re-queues
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_failed_retry_clears_prior_attempt_and_reruns(monkeypatch):
    run = _seed_run(slug="retry-clean")
    _run_real_audit(run, monkeypatch)
    run.refresh_from_db()
    assert run.state == RunState.GATE_1_REVIEW
    first_pages = run.pages.count()
    assert first_pages > 0

    AuditRun.objects.filter(pk=run.pk).update(
        state=RunState.FAILED, error_code="audit_execution_failed"
    )
    run.refresh_from_db()
    outcome = run_website_audit.run(str(run.pk))

    assert outcome.get("idempotent") is not True
    run.refresh_from_db()
    assert run.state == RunState.GATE_1_REVIEW
    assert run.pages.count() == first_pages  # replaced, not duplicated


@pytest.mark.django_db
def test_cancel_during_crawl_is_never_overwritten(monkeypatch):
    from audit_engine import tasks as audit_tasks

    run = _seed_run(slug="cancel-wins")

    class CancellingCrawler:
        def __init__(self, config):
            self.config = config

        def crawl(self, seeds):
            AuditRun.objects.filter(pk=run.pk).update(state=RunState.CANCELLED)
            return CrawlResult(
                pages=(
                    CrawledPage(
                        requested_url="https://example.com.au/",
                        final_url="https://example.com.au/",
                        status_code=200, content_type="text/html",
                        body_sha256="c" * 64, title="Home", meta_description=None,
                        h1=("Welcome",), canonical_url=None, robots_directives=(),
                        links=(), redirect_chain=("https://example.com.au/",),
                    ),
                ),
                failures=(), discovered_count=1, stopped_reason="queue_exhausted",
            )

    monkeypatch.setattr(audit_tasks, "BoundedCrawler", CancellingCrawler)
    outcome = run_website_audit.run(str(run.pk))

    assert outcome.get("cancelled") is True
    run.refresh_from_db()
    assert run.state == RunState.CANCELLED


@pytest.mark.django_db
def test_resume_requeues_a_failed_crawl(monkeypatch, client):
    from audit_engine import tasks as audit_tasks

    run = _seed_run(slug="resume-requeue")
    run.state = RunState.FAILED
    run.error_code = "audit_execution_failed"
    run.save(update_fields=["state", "error_code", "updated_at"])
    RunStage.objects.create(
        run=run, name="collecting", sequence=10, status=StageStatus.FAILED
    )

    kicked: list[str] = []
    monkeypatch.setattr(
        audit_tasks.run_website_audit, "delay", lambda run_id: kicked.append(run_id)
    )

    assert client.login(username="resume-requeue-admin", password="Package-test-password-9911!")  # noqa: S106 - test credential
    response = client.post(f"/runs/{run.pk}/resume/", secure=True)

    assert response.status_code == 302
    assert kicked == [str(run.pk)]
    run.refresh_from_db()
    assert run.state == RunState.FAILED  # the task flips state when it starts


@pytest.mark.django_db
def test_zombie_run_does_not_freeze_backfills(monkeypatch, client, settings):
    from datetime import timedelta

    from django.utils import timezone

    import exporters.tasks as exporter_tasks

    settings.AUTO_BUILD_PACKAGE = True
    run = _seed_run(slug="zombie-gate")
    run.state = RunState.GATE_1_REVIEW
    run.save(update_fields=["state", "updated_at"])

    zombie_project = Project.objects.create(
        client=run.project.client,
        name="Zombie",
        slug="zombie-gate-stuck",
        primary_domain="stuck.example.com.au",
        approved_domains=["stuck.example.com.au"],
        business_type=Project.BusinessType.SERVICE,
    )
    zombie = AuditRun.objects.create(
        project=zombie_project,
        profile="quick",
        idempotency_key="zombie-gate-run",
        rule_version="1.0.0",
        state=RunState.COLLECTING,
    )
    AuditRun.objects.filter(pk=zombie.pk).update(
        updated_at=timezone.now() - timedelta(hours=3)
    )

    queued: list[dict] = []
    monkeypatch.setattr(
        exporter_tasks.build_audit_package,
        "apply_async",
        lambda args=None, **kwargs: queued.append(kwargs),
    )

    assert client.login(username="zombie-gate-admin", password="Package-test-password-9911!")  # noqa: S106 - test credential
    assert client.get("/", secure=True).status_code == 200
    assert queued == [{"queue": "render"}]


def test_verify_tree_allows_kakawa_when_provider_evidence_names_it(tmp_path):
    """A competitor who is also a client is market evidence, not a leak."""
    from exporters.run_package import _verify_tree

    root = tmp_path / "pkg"
    _write(root, tree.SUMMARY_MARKDOWN, b"# Top competitor: kakawachocolates.com.au\n")

    contaminated = {"client": {"name": "ChocolArts"}, "competitors": []}
    failures, _counts, _total = _verify_tree(root, contaminated)
    assert any("Kakawa" in failure for failure in failures)

    evidenced = {
        "client": {"name": "ChocolArts"},
        "competitors": [{"domain": "kakawachocolates.com.au"}],
    }
    failures, _counts, _total = _verify_tree(root, evidenced)
    assert not [failure for failure in failures if "Kakawa" in failure]
